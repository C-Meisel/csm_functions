''' This module contains functions to format, plot, analyze, and save EIS and DRT spectra for multiple 
Electrochemical cell tests.  The data comes from a Gamry Potentiostat
# C-Meisel
'''
# import bayes_drt2
'Imports'
import os
import pandas as pd
import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import natsort
import seaborn as sns
import scipy as scipy
import cmasher as cmr
import sys # Used to go through the system
import traceback # Used to look at errors
from mpl_toolkits.axes_grid1.axes_divider import make_axes_locatable
from mpl_toolkits.axes_grid1.axes_divider import make_axes_locatable
from matplotlib.colors import ListedColormap
import matplotlib.ticker as ticker
from matplotlib.ticker import FuncFormatter

from hybdrt.models import DRT, elements, drtbase
import hybdrt.plotting as hplt
from hybdrt.fileload import read_eis, get_eis_tuple, get_timestamp

from .plotting import plot_peis, plot_peiss, lnpo2
from .fit_drt import dual_drt_save, pfrt_drt_save
from .data_formatting import peis_data, read_dta

' --- Font stuff'
from matplotlib import font_manager
SSP_r = '/Library/Fonts/SourceSansPro-Regular.ttf'
cm_ss = '/Library/Fonts/cmunss.ttf'
ubuntu = '/Library/Fonts/Ubuntu-Regular.ttf'
ubuntu_bold = '/Library/Fonts/Ubuntu-bold.ttf'

font_manager.fontManager.addfont(SSP_r)
font_manager.fontManager.addfont(cm_ss)
font_manager.fontManager.addfont(ubuntu)
font_manager.fontManager.addfont(ubuntu_bold)

def standard_performance(loc:str, jar:str, area:float=0.5, dual_pfrt = True, peaks_to_fit:int = 'best_id',
                          bayes_factors:bool = False, **peis_args):
    '''
    Plots the EIS, dual-fits and plots the DRT, and prints the ohmic and polarization resistance for a cell
    Generally this function is used on the first EIS spectra taken at standard testing conditions

    Parameters
    ----------
    loc, str: (path to a directory)
        The location of the folder containing the EIS files (path to folder)
    jar, str: (path to a directory)
        The location of the jar containing the map-fits (path to jar)
    area, float: 
        The active cell area in cm^2 (default = 0.5)
    dual_pfrt, bool: (default = True)
        If True, the dual regression model and the pfrt model are used to fit the EIS
        If False, the drt-dop model is used to fit the EIS
    peis_args, dict: 
        Any additional arguments to be passed to the plot_peis function
    peaks_to_fit, str/int: (default: 'best_id')
        Sets the number of discrete elements in the EIS to fit in the DRT
        This basically just sets the amout of peaks to fit
        if this is set to the default 'best_id' then it fits the number of peaks suggested by dual_drt
        if this is set to a integer, it fit the number of peaks set by that integer
    bayes_factors, boolean: (default: False)
        If true, plots hte bayes factors from the dual regression
        These show the likelyness that a specific model is a good fit

    Return --> None, but it plots EIS and DRT. It also prints the ohmic and rp values of the cell
    '''
    cell_name = os.path.basename(loc).split("_", 1)[0] #gets the name of the cell

    if dual_pfrt == True:
        # --- Fitting DRT, and making sure I am not re-fitting it if it has already been fit
        pickle_jar = os.listdir(jar)

        # # - Dual fit check - Will place back in later
        # dual_pickle_name = 0
        # dual_fit_file = cell_name + '_standard_dual.pkl'
        # for pickle in pickle_jar: # checks to see if this has already been fit, if so name gets set to 1
        #     if dual_fit_file == pickle:
        #         dual_pickle_name = dual_pickle_name + 1
        #         break

        # if dual_pickle_name == 0:
        #     dual_fit_name = dual_fit_file.removesuffix('_dual.pkl')
        #     dual_drt_save(loc,jar,dual_fit_name)

        # - Dual fit
        dual_drt = DRT() # Create a DRT instance
        dual_drt.fit_dop = True
        df = read_eis(loc)
        freq,z = get_eis_tuple(df) # Get relavent data from EIS dataframe
        dual_drt.dual_fit_eis(freq, z, discrete_kw=dict(prior=True, prior_strength=None),) # Fit the data 
        tau = dual_drt.get_tau_eval(20)

        # - Best Model to use to dual fit the DRT peaks:
        if peaks_to_fit == 'best_id': 
            best_id = dual_drt.get_best_candidate_id('discrete', criterion='lml-bic')
            peaks = best_id

        else: peaks = peaks_to_fit

        if bayes_factors == True:
            # Plot normalized Bayes factors for each model
            # This is taken directly from https://github.com/jdhuang-csm/hybrid-drt/blob/main/tutorials/Fitting%20EIS%20data.ipynb
            # Credit to Dr. Jake Huang
            fig, axes = plt.subplots(1, 3, figsize=(7, 2.25), sharex=True, sharey=True)
            for i, crit in enumerate(['bic', 'lml', 'lml-bic']):
                dual_drt.plot_norm_bayes_factors('discrete', criterion=crit, marker='o', ax=axes[i])
                axes[i].set_title(crit.upper())
            
            if i > 0:
                axes[i].set_ylabel('')
                
            fig.tight_layout()

        # - Setting the Dual Model    
        model_dict = dual_drt.get_candidate(peaks,'discrete')
        model = model_dict['model']

        # - Probability function of relaxation time fit check
        pfrt_pickle_name = 0
        pfrt_fit_file = cell_name+'_standard_pfrt.pkl'
        for pickle in pickle_jar: # checks to see if this has already been fit, if so name gets set to 1
            if pfrt_fit_file == pickle:
                pfrt_pickle_name = pfrt_pickle_name + 1
                break

        if pfrt_pickle_name == 0:
            pfrt_fit_name = pfrt_fit_file.removesuffix('_pfrt.pkl')
            pfrt_drt_save(loc,jar,pfrt_fit_name)

        # ---- Loading DRT and plotting
        # dual_drt = DRT()
        pfrt_drt = DRT()

        # dual_drt.load_attributes(os.path.join(jar,dual_fit_file))
        pfrt_drt.load_attributes(os.path.join(jar,pfrt_fit_file))

        pfrt_basis_tau = pfrt_drt.basis_tau # gaining the basis tau of the pfrt fit
        pf = pfrt_drt.predict_pfrt(pfrt_basis_tau) # saving the pfrt as an object
        
        # -- Plotting
        fig, ax = plt.subplots()
        m_blue = '#21314D'
        m_orange = '#c1741d'
        
        # dual_drt.plot_distribution(c=m_blue, plot_ci=True, label='Dual DRT', return_line=True, ax=ax, area=area, mark_peaks = True)
        # dual_drt.plot_candidate_distribution(best_id, 'discrete', label='Dual DRT',
        #             c=m_blue, ax=ax, area=area)

        model.plot_distribution(tau, ax=ax, area=area,label='Dual DRT', c=m_blue, mark_peaks=True)

        ax2 = ax.twinx()
        ax2.plot(pfrt_basis_tau, pf, label='PFRT', color=m_orange)
        
        # - Formating
        ax2.set_ylabel('$p$',fontsize='xx-large',color=m_orange, labelpad = -10)
        ax.legend(loc='upper left', fontsize = 'x-large', frameon = False, labelcolor = m_blue, handletextpad=0.3)
        ax2.legend(fontsize = 'x-large', frameon = False, labelcolor = m_orange, handletextpad=0.3)

        # - Excessive formatting
        ax.xaxis.label.set_size('xx-large')
        ax.tick_params(axis='x', labelsize='x-large')

        ax.yaxis.label.set_size('xx-large')
        ax.yaxis.label.set_color(m_blue)
        ax.tick_params(axis='y', labelsize='x-large', labelcolor=m_blue, color=m_blue)
        ax.spines['left'].set_color(m_blue) 

        ax2.tick_params(axis='y', labelsize='x-large', labelcolor=m_orange, color=m_orange)
        ax2.spines['right'].set_color(m_orange)
        ax2.set_yticks([0,1])

        ax2.spines['top'].set_visible(False)
        ax.spines['top'].set_visible(False)

        plt.tight_layout()
        plt.show()

        # # ---- Plotting EIS
        ohmic = dual_drt.predict_r_inf()*area
        rp = dual_drt.predict_r_p()*area
        ohmic_rtot = [round(float(ohmic),2),round(float(rp),2)+round(float(ohmic),2)]
        plot_peis(area,loc,**peis_args) #Plots standard PEIS spectra

    else:
        # -- Initializing drt instance and dop model
        drt = DRT()
        drt.fit_dop=True
        
        # -- Gathering data
        df = read_eis(loc)
        freq,z = get_eis_tuple(df) # Get relavent data from EIS dataframe

        # -- Fitting DRT
        drt.fit_eis(freq, z)    
        tau = drt.get_tau_eval(20)
    
        # -- Plotting
        fig, ax = plt.subplots()
        m_blue = '#21314D'
        m_orange = '#c1741d'
        mark_peaks_kw = {'color':m_orange}
        drt.plot_distribution(tau, ax=ax, area=area,label=cell_name,mark_peaks=True,
                           scale_prefix="", c=m_blue, plot_ci=True,mark_peaks_kw=mark_peaks_kw)
                
        # - Adding Frequency scale:
        def Tau_to_Frequency(T):
            return 1 / (2 * np.pi * T)

        freq_ax = ax.secondary_xaxis('top', functions=(Tau_to_Frequency, Tau_to_Frequency))
        freq_ax.set_xlabel('$f$ (Hz)',size='xx-large')
        freq_ax.tick_params(axis='x',labelsize='x-large')


        # - Excessive formatting
        ax.xaxis.label.set_size('xx-large')
        ax.tick_params(axis='both', labelsize='x-large')
        ax.yaxis.label.set_size('xx-large')
        ax.spines['right'].set_visible(False)

        plt.tight_layout()
        plt.show()
        # # ---- Plotting EIS
        ohmic = drt.predict_r_inf()*area
        rp = drt.predict_r_p()*area
        ohmic_rtot = [round(float(ohmic),2),round(float(rp),2)+round(float(ohmic),2)]
        plot_peis(area,loc,**peis_args) #Plots standard PEIS spectra

    print('Standard Ohmic',round(ohmic,3),'\u03A9cm\u00b2')
    print('Standard Rp',round(rp,3),'\u03A9cm\u00b2')

def plot_drtdop(loc:str, area:float, label:str = None, ax:plt.Axes = None, scale_prefix = "",
                 mark_peaks = True, legend = True, print_resistance:bool=False,
                 publication:bool = False, save_plot = None, nonneg=True, **kwargs):
    '''
    Quicker version to DRT-DOP fit and plot the ensuing DRT spectra.
    This function supports multiple graphs stacked on each other
    If one spectra is being plot, the fig and ax are made inside the function.

    Parameters
    ----------
    loc, str: (path to a directory)
        The location of the folder containing the EIS files (path to folder)
    area, float:
        The active cell area in cm^2
    label, str: (default, None)
        The name of the spectra. Will be displayed in the legend
    ax, plt.Axes: (default: None)
        matplotlib axes object. Used to plot the spectra. Needed to plot different spectra on the same figure
    mark_peaks, bool: (default: True)
        Whether or not to mark the peaks
    legend, bool: (default: True)
        Whether or not to show the legend
    print_resistance, bool: (default: False)
        whether or not to print the resistance values of the cell
    publication, bool: (default = False)
        If false the figure is formatted for a presentation
        If true the figure is formatted to be a subfigure in a journal paper.
        Setting publication to true increases all feature sizes
    save_plot, str: (default = None)
        If this is not none, the plot will be saved.
        Save_plot is the file name and path of the saved file.
    nonneg, bool: (default = True)
        no negative values
        If set to true, the DRT is constrained such that all values are positive
        However if the potential for negative DRT is desired (Such as EC mode fitting)
        then set nonneg to False
    
    Return --> DRT instance, and it plots and shows one or more plots
    '''
    ' Plotting params - For Thesis '
    mpl.rcParams['font.family'] = 'Ubuntu'
    spine_thickness = 1.3
    txt_spine_color = '#212121' # '#212121' # 'black' #333333 '#666666'
    plt.rcParams.update({
        'axes.linewidth': spine_thickness,        # Spine thickness
        'xtick.major.width': spine_thickness,     # Major tick thickness for x-axis
        'ytick.major.width': spine_thickness,     # Major tick thickness for y-axis
        'xtick.major.size': spine_thickness * 3,      # Major tick length for x-axis
        'ytick.major.size': spine_thickness * 3,      # Major tick length for y-axis
        
        'text.color': txt_spine_color,                   # Color for all text
        'axes.labelcolor': txt_spine_color,              # Color for axis labels
        'axes.edgecolor': txt_spine_color,               # Color for axis spines
        'xtick.color': txt_spine_color,                  # Color for x-axis tick labels and ticks
        'ytick.color': txt_spine_color,                  # Color for y-axis tick labels and ticks
    })

    # -- Initializing drt instance and dop model
    drt = DRT()
    drt.fit_dop=True
    
    # -- Gathering data
    df = read_eis(loc)
    freq,z = get_eis_tuple(df) # Get relavent data from EIS dataframe

    # -- Fitting DRT
    drt.fit_eis(freq, z, nonneg=nonneg)    
    tau = drt.get_tau_eval(20)

    # -- Plotting
    solo = None 
    plot_ci = False
    # c = None
    mark_peaks_kw = {'sizes':[75]}

    if ax == None: # If only one spectra is being plot, create the fig and axis in the function
        solo = True
        plot_ci = True
        c = '#21314D' # mines blue
        m_orange = '#c1741d'
        mark_peaks_kw = {'color':m_orange,'sizes':[75]}
    
        fig, ax = plt.subplots()

    drt.plot_distribution(tau, ax=ax, area=area,label=label,mark_peaks=mark_peaks,
                           scale_prefix=scale_prefix, plot_ci=plot_ci,
                            mark_peaks_kw=mark_peaks_kw, **kwargs) # c = c

    # - Adding Frequency scale:
    def Tau_to_Frequency(T):
        return 1 / (2 * np.pi * T)

    # freq_ax = ax.secondary_xaxis('top', functions=(Tau_to_Frequency, Tau_to_Frequency))

    # - Formatting
    if legend == True:
        ax.legend(fontsize='x-large')

    # - Excessive formatting
    if publication == False:
        label_size = 23
        tick_size = label_size * 0.85
        ax.xaxis.label.set_size(label_size) 
        ax.yaxis.label.set_size(label_size)
        ax.tick_params(axis='both', labelsize = tick_size)

        # - Frequency Ax
        freq_ax = ax.secondary_xaxis('top', functions=(Tau_to_Frequency, Tau_to_Frequency))
        freq_ax.set_xlabel('$f$ (Hz)',size = label_size,labelpad=10)
        freq_ax.tick_params(axis='x',labelsize= tick_size)

    if publication == True:
        label_size = 28
        tick_size = label_size * 0.80
        spine_width = 2 # 
        # - Main Axes
        x_ticks = np.array([1e-7,1e-5,1e-3,1e-1,1e1])
        ax.set_xticks(x_ticks)
        ax.yaxis.set_major_formatter(FuncFormatter(lambda y, _: f'{y*1000:.0f}'))
        ax.xaxis.set_major_formatter(FuncFormatter(lambda x, _: f'{int(np.log10(x))}'))
        ax.set_ylabel(r'$\gamma$ (m$\Omega$ $\cdot$ cm$^2$)',fontsize=label_size)
        ax.set_xlabel(r'log$_{10}\tau$',fontsize=label_size)
        ax.tick_params(axis='both', labelsize=tick_size,width=2,length=6)
        ax.spines['left'].set_linewidth(spine_width)
        ax.spines['bottom'].set_linewidth(spine_width)
        # - Frequency Ax
        freq_ax = ax.secondary_xaxis('top', functions=(Tau_to_Frequency, Tau_to_Frequency))
        freq_x_ticks = np.array([1e6,1e4,1e2,1e0,1e-2])
        freq_ax.set_xticks(freq_x_ticks)
        freq_ax.set_xlabel(r'$f$ (log$_{10}$Hz)',size=label_size,labelpad = 10)
        freq_ax.tick_params(axis='x',labelsize=tick_size,width=2,length=6)
        freq_ax.spines['top'].set_linewidth(spine_width)
        freq_ax.xaxis.set_major_formatter(FuncFormatter(lambda x, _: f'{int(np.log10(x))}'))



    ax.spines['right'].set_visible(False)
    ax.yaxis.set_major_locator(ticker.MaxNLocator(nbins=4))


    # - Printing resistance values if desired
    if print_resistance == True:
        ohmic = round(drt.predict_r_inf() * area,3)
        rp = round(drt.predict_r_p() * area,3)
        cell_name = os.path.basename(loc)

        print(cell_name + ' Ohmic: ',ohmic,'\u03A9cm\u00b2')
        print(cell_name + ' Rp: ',rp,'\u03A9cm\u00b2')

    if solo == True:
        plt.tight_layout()
        
        if save_plot is not None:
            fmat = save_plot.split('.', 1)[-1]
            fig.savefig(save_plot, dpi=300, format=fmat, bbox_inches='tight')
       
        plt.show()

    return(drt)

def plot_dop(loc:str, label:str = None, scale_prefix = "",
                legend = True, **kwargs):
    '''
    Quicker version to use DRT-DOP to fit and plot the ensuing DOP spectra only.

    Parameters
    ----------
    loc, str: (path to a directory)
        The location of the folder containing the EIS files (path to folder)
    label, str: (default, None)
        The name of the spectra. Will be displayed in the legend
    ax, plt.Axes: (default: None)
        matplotlib axes object. Used to plot the spectra. Needed to plot different spectra on the same figure
    legend, bool: (default: True)
        Whether or not to show the legend
    Return --> DRT instance, and it plots and shows one or more plots
    '''
    # -- Initializing drt instance and dop model
    drt = DRT()
    drt.fit_dop=True
    
    # -- Gathering data
    df = read_eis(loc)
    freq,z = get_eis_tuple(df) # Get relavent data from EIS dataframe

    # -- Fitting DRT
    drt.fit_eis(freq, z)    
    tau = drt.get_tau_eval(20)

    # -- Plotting
    drt.plot_dop(normalize=True,normalize_tau=(tau.min(), tau.max()) ,label=label, **kwargs) # 
    plt.legend()
    plt.tight_layout()
    plt.show()

    return(drt)

def quick_dualdrt_plot(loc:str, area:float, label:str = None, ax:plt.Axes = None, peaks_to_fit:int = 'best_id', scale_prefix = "",
                       mark_peaks = True, legend = True, **kwargs):
    '''
    Quicker version to fit and plot a dual DRT spectra to EIS data.
    This function supports multiple graphs stacked on each other
    If one spectra is being plot, the fig and ax are made inside the function.

    Parameters
    ----------
    loc, str: (path to a directory)
        The location of the folder containing the EIS files (path to folder)
    area, float:
        The active cell area in cm^2
    label, str: (default, None)
        The name of the spectra. Will be displayed in the legend
    ax, plt.Axes:
        matplotlib axes object. Used to plot the spectra. Needed to plot different spectra on the same figure
    peaks_to_fit, str/int: (default: 'best_id')
        Sets the number of discrete elements in the EIS to fit in the DRT
        This basically just sets the amout of peaks to fit
        if this is set to the default 'best_id' then it fits the number of peaks suggested by dual_drt
        if this is set to a integer, it fit the number of peaks set by that integer
    mark_peaks, bool: (default: True)
        Whether or not to mark the peaks
    legend, bool: (default: True)
        Whether or not to show the legend

    Return --> DRT instance, and it plots and shows one or more plots
    '''
    # -- Gathering data
    drt = DRT()
    df = read_eis(loc)
    freq,z = get_eis_tuple(df) # Get relavent data from EIS dataframe
    drt.dual_fit_eis(freq, z, discrete_kw=dict(prior=True, prior_strength=None)) # Fit the data prior_strength=None
    tau = drt.get_tau_eval(20)

    # -- Selecting the number of peaks for the drt distribution to have.
    if peaks_to_fit == 'best_id':
        best_id = drt.get_best_candidate_id('discrete', criterion='lml-bic') #lml-bic'
        peaks = best_id

    else: peaks = peaks_to_fit

    # - Selecting the model. The number of peaks to plot and fit
    model_dict = drt.get_candidate(peaks,'discrete')
    model = model_dict['model']
    
    # --- Plotting
    solo = None 
    if ax == None: # If only one spectra is being plot, create the fig and axis in the function
        solo = True
        fig, ax = plt.subplots()

    model.plot_distribution(tau, ax=ax, area=area,label=label,mark_peaks=mark_peaks, scale_prefix=scale_prefix, **kwargs)

    # - Formatting
    if legend == True:
        ax.legend(fontsize='x-large')
    ax.xaxis.label.set_size('xx-large') 
    ax.yaxis.label.set_size('xx-large')
    ax.tick_params(axis='both', labelsize='x-large')

    # - Excessive formatting
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    if solo == True:
        plt.tight_layout()
        plt.show()

    return(drt)

def po2_plots_dual(folder_loc:str, fit_path:str, area:float, eis:bool=True, drt:bool=True,
                 o2_dependence:bool=True, drt_peaks:bool=True, print_resistance_values:bool=False,
                  ncol:int=1, legend_loc:str='best', flow100:bool = False, flow200:bool = False, cut_inductance:bool = False,
                  overwrite:bool = False, peaks_to_fit:int = 'best_id', drt_model:str = 'dual',
                  save_eis:str=None,save_drt:str=None):
    '''
    Searches through the folder_loc for all the changes in O2 concentration EIS files.
    Plots the EIS for each concentration in one plot if eis=True and the DRT of each concentration in another if drt=True
    dual-fits all of the eis files.
    If O2_dependence = True, plots the ln(1/ASR) vs. ln(O2 concentration) for the ohmic and polarization resistance
    This funciton uses dual_drt regression fitting.

    Parameters:
    -----------
    folder_loc, str: (path to folder)
        The location of the folder containing the EIS files 
    fit_path, str: (path to folder)
        The location of the folder to save the map-fits
    area, float:
        The active cell area in cm^2
    eis, bool: (default = True)
        If True, plots the EIS for each concentration in one plot
    drt, bool: (default = True)
        If True, plots the DRT for each concentration in another plot
    o2_dependence, bool: (default = True)
        If True, plots the ln(1/ASR) vs. ln(O2 concentration) for the ohmic and polarization resistance
    drt_peaks, bool: (default = True)
        If True, the DRT peaks are fit and plotted. If the peaks have not been fit, they
        will be fit and added the the cell data excel sheet. If the data exists in the spreadsheet already
        then that data will be plotted.
    print_resistance_values, bool: (default = False)
        If True, prints the ohmic and polarization resistance for each concentration
    ncol, int: (default = 1)
        The number of columns to use for the plot legend 
    legend_loc, str: 
        The location of the legend (default = 'best'). The other option is to put 
        the legend outside the figure and this is done by setting legend_loc = 'outside'
    flow100, bool: (default = False)
        Set this to true if the total flow rate for the PO2 test was 100 SCCM
        for my older tests (cell 16 and below) the total flow rate was 50 SCCM
        This just changes the list of strings to look for
    cut_inductance, bool: (default = False)
        If this is set to true, the negative inductance values at the beginning of the DataFrame
    overwrite, bool: (default = False)
        If the data already exists in the cell excel file, then this will overwrite that data with
        the data currently being fit when overwrite=True.
    peaks_to_fit, str/int: (default: 'best_id')
        Sets the number of discrete elements in the EIS to fit in the DRT
        This basically just sets the amout of peaks to fit
        if this is set to the default 'best_id' then it fits the number of peaks suggested by dual_drt
        if this is set to a integer, it fit the number of peaks set by that integer
    drt_model, str: (default = dual)
        which type of drt used to analyze the data
        if drt = 'dual' the dual regression model is used to fit the data
        if drt = 'drtdop' then the drtdop model is used to fit the data
    save_eis, str: (default = None)
        If this is not none, the EIS figure will be saved.
        Save_eis is the file name and path of the saved file.
    save_drt, str: (default = None)
        If this is not none, , the DRT figure will be saved.
        Save_drt is the file name and path of the saved file.

    Return --> None, but it plots and shows one or more plots
    '''
    
    # +++++++++----- Finding and Dual DRT fitting the PO2 EIS
    dta_files = [file for file in os.listdir(folder_loc) if file.endswith('.DTA')] #Makes a list of all .DTA files in the folder loc

    if flow100 == True: # setting variables to look for
        substring_list = ['PEIS_20O2.80Ar','PEIS_40O2.60Ar','PEIS_60O2.40Ar','PEIS_80O2.20Ar','PEIS_100O2.0Ar']
    elif flow200 == True:
        substring_list = ['PEIS_40O2.160Ar','PEIS_80O2.120Ar','PEIS_120O2.80Ar','PEIS_160O2.40Ar','PEIS_200O2.0Ar']
    else:
        substring_list = ['PEIS_10O2.40Ar','PEIS_20O2.30Ar','PEIS_30O2.20Ar','PEIS_40O2.10Ar','PEIS_50O2.0Ar']
        # substring_list = ['10May21.1_PEIS_50air.3steam_50H2.0Ar_550C_2.DTA','PEIS_20O2.30Ar','PEIS_30O2.20Ar','PEIS_40O2.10Ar','PEIS_50O2.0Ar']

    O2_conc_list = [dta_files for dta_files in dta_files if any(sub in dta_files for sub in substring_list)] # placing all the changes in O2 eis files into a list
    cell_name = os.path.basename(fit_path).split("_", 2)[1]

    O2_conc_list = sorted(O2_conc_list, key=lambda x: int((x[x.find('PEIS_')+len('PEIS_'):x.rfind('O2')]))) #Sorts numerically by PO2
    
    ' ---------- Plotting all the PO2 concentration EIS spectra'
    if eis == True:
        plt.figure() # Initializing the Figure

        for peis in O2_conc_list:
            # --- Finding and formatting
            loc = os.path.join(folder_loc, peis) # creates the full path to the file
            po2 = peis[peis.find('PEIS_')+len('PEIS_'):peis.rfind('O2')] #gets the pO2 from the file name

            if flow200 == True:
                po2_int = int(po2)
                po2 = str(int(po2_int / 2))

            elif flow100 == False and flow200 == False:
                po2_int = int(po2)
                po2 = str(po2_int * 2)
                nyquist_name = po2 + '% O$_2$'
            else:
                nyquist_name =  po2 + '% O$_2$'


            # --- Plotting
            plot_peiss(area,nyquist_name,loc,ncol=ncol,legend_loc=legend_loc,cut_inductance = cut_inductance)
        
        # - Saving the figure
        if save_eis is not None:
            fmat_eis = save_eis.split('.', 1)[-1]
            plt.savefig(save_eis, dpi=300, format=fmat_eis, bbox_inches='tight') 

        plt.show()

    ' ---------- inverting, plotting, and dual fitting all the PO2 concentration DRT spectra'
    if drt == True: 
        fig, ax = plt.subplots() #initializing plots for DRT
        # --- Initializing lists for further analysis
        O2_conc = np.array([]) # Concentration of Oxygen
        ohmic_asr = np.array([]) # ohm*cm^2
        rp_asr = np.array([]) # ohm*cm^2
        df_tau_r = pd.DataFrame(columns = ['O2 Concentration (%)','Tau','Resistance']) #Initializing DataFrame to save DRT peak data

        for peis in O2_conc_list: #For loop to plot the DRT of all PO2 files (probably a more elegant way, but this works)
            # --- Finding and formatting
            loc = os.path.join(folder_loc, peis) # creates the full path to the file
            po2 = peis[peis.find('PEIS_')+len('PEIS_'):peis.rfind('O2')] #gets the pO2 from the file name
            
            if flow200 == True:
                po2_int = int(po2)
                po2 = str(int(po2_int / 2))

            elif flow100 == False and flow200 == False:
                po2_int = int(po2)
                po2 = str(po2_int * 2)
                nyquist_name = po2 + '% O$_2$'

            else:
                nyquist_name =  po2 + '% O$_2$'

            # if flow100 == True:
            #     po2_int = int(po2)
            #     po2 = str(po2_int * 2)
            #     nyquist_name = po2 + '% O$_2$'

            # --- Inverting the EIS data
            drt = DRT()
            df = read_eis(loc)
            freq,z = get_eis_tuple(df) # Get relavent data from EIS dataframe
            label = po2 + '% O$_2$'
            
            if drt_model == 'dual':
                drt.dual_fit_eis(freq, z, discrete_kw=dict(prior=True, prior_strength=None)) # Fit the data
                tau = drt.get_tau_eval(20)

                # - Selecting the number of peaks for the drt distribution to have.
                if peaks_to_fit == 'best_id':
                    best_id = drt.get_best_candidate_id('discrete', criterion='lml-bic')
                    peaks = best_id

                else: 
                    peaks = peaks_to_fit
            
                # - Selecting the model. The number of peaks to plot and fit
                model_dict = drt.get_candidate(peaks,'discrete')
                model = model_dict['model']

                # --- Plotting
                model.plot_distribution(tau, ax=ax, area=area,label=label,mark_peaks=True)

            elif drt_model == 'drtdop':
                drt.fit_dop=True
                drt.fit_eis(freq, z)
                tau = drt.get_tau_eval(20)

                # - Plotting
                mark_peaks_kw = {'sizes':[75]}
                kwargs = {
                    'linewidth': 2,
                }
                drt.plot_distribution(tau, ax=ax, area=area,label=label,mark_peaks=True,
                                    scale_prefix="", plot_ci=False,mark_peaks_kw=mark_peaks_kw,
                                    **kwargs)
            else:
                print('In order to plot DRT, it must be fit with the dual or the drtdop model') 
                print('Set drt= \'dual\' or to \'drtdop\'') 

            # --- Appending resistance values to lists for lnPO2 plotting
            O2_conc = np.append(O2_conc,float(po2)/100)
            ohmic_asr = np.append(ohmic_asr,drt.predict_r_inf() * area)
            rp_asr = np.append(rp_asr,drt.predict_r_p() * area)
            append_drt_peaks(df_tau_r, drt, area, po2, peaks_to_fit = peaks_to_fit,
                             drt_model = drt_model)

            # --- Printing the resistance values if that is desired
            if print_resistance_values == True:
                print(po2 + '% Oxygen Ohmic', round(drt.predict_r_inf() * area,3), '\u03A9 cm$^2^')
                print(po2 + '% Oxygen Rp', round(drt.predict_r_p() * area,3), '\u03A9 cm$^2^')
        
        # --- Formatting
        # - Adding Frequency scale:
        def Tau_to_Frequency(T):
            return 1 / (2 * np.pi * T)

        freq_ax = ax.secondary_xaxis('top', functions=(Tau_to_Frequency, Tau_to_Frequency))
        
        # - Excessive formatting
        ax_title_size = 23
        tick_label_size = ax_title_size * 0.8
        legend_txt_size = ax_title_size * 0.7

        ax.legend(fontsize=legend_txt_size,frameon=False,handletextpad=0.5,
                  handlelength=1)
        ax.xaxis.label.set_size(ax_title_size) 
        ax.yaxis.label.set_size(ax_title_size)
        ax.tick_params(axis='both', labelsize=tick_label_size)
        ax.spines['right'].set_visible(False)

        freq_ax.set_xlabel('$f$ (Hz)',size=ax_title_size)
        freq_ax.tick_params(axis='x',labelsize=tick_label_size)
    
        plt.tight_layout()

        if save_drt is not None:
            fmat_drt = save_drt.split('.', 1)[-1]
            fig.savefig(save_drt, dpi=300, format=fmat_drt, bbox_inches='tight') 

        plt.show()

        # >>>>>>>>>>>> Creating DataFrames and adding excel data sheets (or creating the file if need be)'
        # ------- Cell Resistance Data
        excel_name = '_' + cell_name + '_Data.xlsx'
        excel_file = os.path.join(folder_loc,excel_name)
        sheet_name = 'pO2_' + drt_model
        exists = False

        exists, writer = excel_datasheet_exists(excel_file,sheet_name)

        if exists == False:
            df_po2 = pd.DataFrame(list(zip(O2_conc*100,ohmic_asr,rp_asr)),
                columns =['O2 Concentration (%)','Ohmic ASR (ohm*cm$^2$)', 'Rp ASR (ohm*cm$^2$)'])
            df_po2.to_excel(writer, sheet_name=sheet_name, index=False) # Writes this DataFrame to a specific worksheet
            writer.close() # Close the Pandas Excel writer and output the Excel file.

        elif exists == True and overwrite == True:
            df_po2 = pd.DataFrame(list(zip(O2_conc*100,ohmic_asr,rp_asr)),
                columns =['O2 Concentration (%)','Ohmic ASR (ohm*cm$^2$)', 'Rp ASR (ohm*cm$^2$)'])
            
            book = load_workbook(excel_file)
            writer = pd.ExcelWriter(excel_file,engine='openpyxl',mode='a',if_sheet_exists='replace') #Creates a Pandas Excel writer using openpyxl as the engine in append mode
            df_po2.to_excel(writer, sheet_name=sheet_name, index=False) # Extract data to an excel sheet
            writer.close() # Close the Pandas Excel writer and output the Excel file.

            df_po2 = pd.read_excel(excel_file,sheet_name)

        # ------- Appending the Peak_fit data to an excel file
        excel_name = '_' + cell_name + '_Data.xlsx'
        excel_file = os.path.join(folder_loc,excel_name)
        peak_data_sheet = 'pO2_' + drt_model + '_DRT_peaks'
        exists_peaks = False

        exists_peaks, writer_peaks = excel_datasheet_exists(excel_file,peak_data_sheet)
    
        if exists_peaks == False: # Make the excel data list
            df_tau_r.to_excel(writer_peaks, sheet_name=peak_data_sheet, index=False) # Extract data to an excel sheet
            writer_peaks.close() # Close the Pandas Excel writer and output the Excel file.
            df_tau_r = pd.read_excel(excel_file,peak_data_sheet)

        elif exists_peaks == True and overwrite == False: #load the data into a DataFrame
            df_tau_r = pd.read_excel(excel_file,peak_data_sheet)

        elif exists_peaks == True and overwrite == True:
            book = load_workbook(excel_file)
            writer_peaks = pd.ExcelWriter(excel_file,engine='openpyxl',mode='a',if_sheet_exists='replace') #Creates a Pandas Excel writer using openpyxl as the engine in append mode
            df_tau_r.to_excel(writer_peaks, sheet_name=peak_data_sheet, index=False) # Extract data to an excel sheet
            writer_peaks.close() # Close the Pandas Excel writer and output the Excel file.

            df_tau_r = pd.read_excel(excel_file,peak_data_sheet)


    ' ------ Plotting the Oxygen dependence'
    if o2_dependence == True:
        lnpo2(ohmic_asr,rp_asr,O2_conc)
    elif drt==False and o2_dependence==True or print_resistance_values==True:
        print('Set drt to True. The cell resistance values are found using Jakes DRT package and thus the EIS spectra need to be fit')

    ' --- DRT peak plotting --- '
    if drt_peaks == True:
        # ----- plotting
        palette = sns.color_palette('crest_r', as_cmap=True)
        plot = sns.scatterplot(x = 'Tau', y = 'Resistance', data = df_tau_r, hue='O2 Concentration (%)',palette = palette,s=69)

        # ----- Ascetics stuff
        sns.set_context("talk")
        fontsize = 14
        sns.despine()
        plot.set_ylabel('ASR (\u03A9 cm$^2$)',fontsize=fontsize)
        plot.set_xlabel('Time Constant (\u03C4/s)',fontsize=fontsize)
        plot.set(xscale='log')

        plt.tight_layout()
        plt.show()

def po2_plots_save(folder_loc:str, fit_path:str, area:float, eis:bool=True, drt:bool=True,
                 o2_dependence:bool=True, drt_peaks:bool=True, print_resistance_values:bool=False,
                  ncol:int=1, legend_loc:str='best', flow100:bool = False, cut_inductance:bool = False,
                  overwrite = False):
    '''
    Searches through the folder_loc for all the changes in O2 concentration EIS files.
    PFRT-fits all of hte eis files and saves them to fit_path. If the files are already fit, they will not be re-fit.
    Plots the EIS for each concentration in one plot if eis=True and the DRT of each concentration in another if drt=True
    If O2_dependence = True, plots the ln(1/ASR) vs. ln(O2 concentration) for the ohmic and polarization resistance

    Parameters:
    -----------
    folder_loc, str: (path to folder)
        The location of the folder containing the EIS files 
    fit_path, str: (path to folder)
        The location of the folder to save the map-fits
    area, float:
        The active cell area in cm^2
    eis, bool: (default = True)
        If True, plots the EIS for each concentration in one plot
    drt, bool: (default = True)
        If True, plots the DRT for each concentration in another plot
    o2_dependence, bool: (default = True)
        If True, plots the ln(1/ASR) vs. ln(O2 concentration) for the ohmic and polarization resistance
    drt_peaks, bool: (default = True)
        If True, the DRT peaks are fit and plotted. If the peaks have not been fit, they
        will be fit and added the the cell data excel sheet. If the data exists in the spreadsheet already
        then that data will be plotted.
    print_resistance_values, bool: (default = False)
        If True, prints the ohmic and polarization resistance for each concentration
    ncol, int: (default = 1)
        The number of columns to use for the plot legend 
    legend_loc, str: 
        The location of the legend (default = 'best'). The other option is to put 
        the legend outside the figure and this is done by setting legend_loc = 'outside'
    flow100, bool: (default = False)
        Set this to true if the total flow rate for the PO2 test was 100 SCCM
        for my older tests (cell 16 and below) my total flow rate was 100 SCCM
        This just changes the list of strings to look for
    cut_inductance, bool: (default = False)
        If this is set to true, the negative inductance values at the beginning of the DataFrame
    overwrite, bool: (default = False)
        If the data already exists in the cell excel file, then this will overwrite that data with
        the data currently being fit when overwrite=True.
        
    Return --> None, but it plots and shows one or more plots
    '''
    
    # +++++++++----- Finding and Dual DRT fitting the PO2 EIS
    dta_files = [file for file in os.listdir(folder_loc) if file.endswith('.DTA')] #Makes a list of all .DTA files in the folder loc
    substring_list = ['PEIS_20O2.80Ar','PEIS_40O2.60Ar','PEIS_60O2.40Ar','PEIS_80O2.20Ar','PEIS_100O2.0Ar'] # setting variables to look for
    O2_conc_list = [dta_files for dta_files in dta_files if any(sub in dta_files for sub in substring_list)] # placing all the changes in O2 eis files into a list
    cell_name = os.path.basename(fit_path).split("_", 2)[1]
    
    for peis in O2_conc_list:
        pickle_jar = os.listdir(fit_path)

        loc = os.path.join(folder_loc, peis) # creates the full path to the file
        po2 = peis[peis.find('PEIS_')+len('PEIS_'):peis.rfind('O2')] #gets the pO2 from the file name
        fit_name = cell_name + '_' + po2 + 'pO2'

        # +++++ checks to see if this has already been fit, if so then the eis will no be re-fit
        pickle_name = 0
        for pickle in pickle_jar: 
            file_name = fit_name + '_pfrt.pkl'
            if file_name == pickle:
                pickle_name = pickle_name + 1
                break
            
        if pickle_name == 0:
            pfrt_drt_save(loc,fit_path,fit_name,indicator=True)

    if flow100 == True: # Used for older cell data taken at 100 SCCM
        substring_list = ['PEIS_10O2.40Ar','PEIS_20O2.30Ar','PEIS_30O2.20Ar','PEIS_40O2.10Ar','PEIS_50O2.0Ar']

    O2_conc_list = sorted(O2_conc_list, key=lambda x: int((x[x.find('PEIS_')+len('PEIS_'):x.rfind('O2')]))) #Sorts numerically by PO2
    
    ' ---------- Plotting all the PO2 concentration EIS spectra'
    if eis == True:
        plt.figure() # Initializing the Figure

        for peis in O2_conc_list:
            # --- Finding and formatting
            loc = os.path.join(folder_loc, peis) # creates the full path to the file
            po2 = peis[peis.find('PEIS_')+len('PEIS_'):peis.rfind('O2')] #gets the pO2 from the file name

            nyquist_name =  po2 + '% O$_2$'

            if flow100 == True:
                po2_int = int(po2)
                po2 = str(po2_int * 2)
                nyquist_name = po2 + '% O$_2$'

            # --- Plotting
            plot_peiss(area,nyquist_name,loc,ncol=ncol,legend_loc=legend_loc,cut_inductance = cut_inductance)

        plt.show()

    ' ---------- Plotting all the PO2 concentration EIS spectra'
    if drt == True: 
        fig, ax = plt.subplots() #initializing plots for DRT

        # --- Initializing lists for LnPO2
        O2_conc = np.array([]) # Concentration of Oxygen
        ohmic_asr = np.array([]) # ohm*cm^2
        rp_asr = np.array([]) # ohm*cm^2

        for peis in O2_conc_list: #For loop to plot the DRT of all PO2 files (probably a more elegant way, but this works)
            # --- Finding and formatting
            loc = os.path.join(folder_loc, peis) # creates the full path to the file
            po2 = peis[peis.find('PEIS_')+len('PEIS_'):peis.rfind('O2')] #gets the pO2 from the file name

            nyquist_name =  po2 + '% O$_2$'

            if flow100 == True:
                po2_int = int(po2)
                po2 = str(po2_int * 2)
                nyquist_name = po2 + '% O$_2$'

            # --- Plotting
            drt = DRT()

            fit_name = cell_name + '_' + po2 + 'pO2_pfrt.pkl'
            drt.load_attributes(os.path.join(fit_path,fit_name))

            label = po2+ '% O$_2$'
            drt.plot_distribution(label=label, return_line=True, ax=ax,area=area)

            # --- Appending resistance values to lists for lnPO2 plotting
            O2_conc = np.append(O2_conc,float(po2)/100)
            ohmic_asr = np.append(ohmic_asr,drt.predict_r_inf() * area)
            rp_asr = np.append(rp_asr,drt.predict_r_p() * area)

            # --- Printing the resistance values if that is desired
            if print_resistance_values == True:
                print(po2 + '% Oxygen Ohmic', round(drt.predict_r_inf() * area,3), '\u03A9 cm$^2^')
                print(po2 + '% Oxygen Rp', round(drt.predict_r_p() * area,3), '\u03A9 cm$^2^')
        
        # - Formatting
        ax.legend(fontsize='x-large')
        ax.xaxis.label.set_size('xx-large') 
        ax.yaxis.label.set_size('xx-large')
        ax.tick_params(axis='both', labelsize='x-large')
        
        plt.tight_layout()
        plt.show()

        'Creating a DataFrame and adding an excel data sheet (or creating the file if need be)'
        excel_name = '_' + cell_name + '_Data.xlsx'
        excel_file = os.path.join(folder_loc,excel_name)
        sheet_name = 'pO2_pfrt'
        exists = False

        exists, writer = excel_datasheet_exists(excel_file,sheet_name)

        if exists == False:
            df_po2 = pd.DataFrame(list(zip(O2_conc*100,ohmic_asr,rp_asr)),
                columns =['O2 Concentration (%)','Ohmic ASR (ohm*cm$^2$)', 'Rp ASR (ohm*cm$^2$)'])
            df_po2.to_excel(writer, sheet_name=sheet_name, index=False) # Writes this DataFrame to a specific worksheet
            writer.close() # Close the Pandas Excel writer and output the Excel file.

        elif exists == True and overwrite == True:
            df_po2 = pd.DataFrame(list(zip(O2_conc*100,ohmic_asr,rp_asr)),
                columns =['O2 Concentration (%)','Ohmic ASR (ohm*cm$^2$)', 'Rp ASR (ohm*cm$^2$)'])
            
            book = load_workbook(excel_file)
            writer = pd.ExcelWriter(excel_file,engine='openpyxl',mode='a',if_sheet_exists='replace') #Creates a Pandas Excel writer using openpyxl as the engine in append mode
            df_po2.to_excel(writer, sheet_name=sheet_name, index=False) # Extract data to an excel sheet
            writer.close() # Close the Pandas Excel writer and output the Excel file.

            df_po2 = pd.read_excel(excel_file,sheet_name)


    ' ------ Plotting the Oxygen dependence'
    if o2_dependence == True:
        lnpo2(ohmic_asr,rp_asr,O2_conc)
    elif drt==False and o2_dependence==True or print_resistance_values==True:
        print('Set drt to True. The cell resistance values are found using Jakes DRT package and thus the EIS spectra need to be fit')

    ' --- DRT peak fitting and plotting --- '
    if drt_peaks == True:
        o2_map_fits = [file for file in os.listdir(fit_path) if file.find('pO2')!=-1 and file.find('pfrt')!=-1] # gets all fuel cell map fits
        o2_map_fits = natsort.humansorted(o2_map_fits,key=lambda y: (len(y),y)) #sorts by bias
        
        # --- Checking to see if the peaks have already been fit:
        peak_data = False
        excel_name = '_' + cell_name + '_Data.xlsx'
        excel_file = os.path.join(folder_loc,excel_name)
        peak_data_sheet = 'pO2_PFRT_DRT_peak_fits'

        exists_peaks, writer_peaks = excel_datasheet_exists(excel_file,peak_data_sheet)
        
        if exists_peaks == False: # Make the excel data list
            # --- Fitting peaks and appending to a DataFrame
            df_tau_r = pd.DataFrame(columns = ['O2 Concentration (%)','Tau','Resistance']) #Initializing DataFrame to save temperature

            for fit in o2_map_fits: #Loading DRT, fitting peaks, and saving to a DataFrame
                # creating DRT object and loading fits
                drt = DRT()
                drt.load_attributes(os.path.join(fit_path,fit))


                # --- obtain time constants from inverters
                tau = drt.find_peaks() # τ/s
                r = np.array(drt.quantify_peaks()) * area # Ω*cm2

                # --- Obtaining pO2
                o2_conc = int(fit[fit.find(cell_name+'_')+len(cell_name+'_'):fit.rfind('pO2')])

                # Appending tau and r for each peak into df_tau_r
                i = 0
                for τ in tau:
                    df_tau_r.loc[len(df_tau_r.index)] = [o2_conc, τ, r[i]]
                    i = i+1

            df_tau_r.to_excel(writer_peaks, sheet_name=peak_data_sheet, index=False) # Extract data to an excel sheet
            writer_peaks.close() # Close the Pandas Excel writer and output the Excel file.

        elif exists == True and overwrite == False: #load the data into a DataFrame
            df_tau_r = pd.read_excel(excel_file,peak_data_sheet)

        elif exists_peaks == True and overwrite == True:
            # --- Fitting peaks and appending to a DataFrame
            df_tau_r = pd.DataFrame(columns = ['O2 Concentration (%)','Tau','Resistance']) #Initializing DataFrame to save temperature

            for fit in o2_map_fits: #Loading DRT, fitting peaks, and saving to a DataFrame
                # creating DRT object and loading fits
                drt = DRT()
                drt.load_attributes(os.path.join(fit_path,fit))


                # --- obtain time constants from inverters
                tau = drt.find_peaks() # τ/s
                r = np.array(drt.quantify_peaks()) * area # Ω*cm2

                # --- Obtaining pO2
                o2_conc = int(fit[fit.find(cell_name+'_')+len(cell_name+'_'):fit.rfind('pO2')])

                # Appending tau and r for each peak into df_tau_r
                i = 0
                for τ in tau:
                    df_tau_r.loc[len(df_tau_r.index)] = [o2_conc, τ, r[i]]
                    i = i+1
            
            book = load_workbook(excel_file)
            writer_peaks = pd.ExcelWriter(excel_file,engine='openpyxl',mode='a',if_sheet_exists='replace') #Creates a Pandas Excel writer using openpyxl as the engine in append mode
            df_tau_r.to_excel(writer_peaks, sheet_name=peak_data_sheet, index=False) # Extract data to an excel sheet
            writer_peaks.close() # Close the Pandas Excel writer and output the Excel file.

            df_tau_r = pd.read_excel(excel_file,peak_data_sheet)
        
        # ----- plotting
        palette = sns.color_palette('crest_r', as_cmap=True)
        plot = sns.scatterplot(x = 'Tau', y = 'Resistance', data = df_tau_r, hue='O2 Concentration (%)',palette = palette,s=69)

        # ----- Aesthetics stuff
        sns.set_context("talk")
        fontsize = 14
        sns.despine()
        plot.set_ylabel('ASR (\u03A9 cm$^2$)',fontsize=fontsize)
        plot.set_xlabel('Time Constant (\u03C4/s)',fontsize=fontsize)
        plot.set(xscale='log')
        plt.tight_layout()
        plt.show()

def fc_bias_plots_dual(folder_loc:str, area:float, eis:bool=True, drt:bool=True,
                        drt_peaks:bool=True, ncol:int=1, legend_loc:str='best',  
                        overwrite:bool = False, peaks_to_fit:int = 'best_id',
                          drt_model:str = 'dual',save_eis:str=None,save_drt:str=None):
    '''
    Finds all EIS files in the folder_loc taken during bias testing in Fuel Cell mode and plots the EIS
    The corresponding DRT fits are located and also plotted if drt = True
    If drt_peaks = True, the DRT spectra are fit, and the resistance of each DRT peak at each condition is plotted
    All data is append to a sheet in the cell data excel file if it does not already exist. If the data
    does already exist that data is called upon for plotting

    The .DTA EIS files are taken during a Gamry sequence that I use for testing the cell performance under
    various biases

    Parameters:
    -----------
    folder_loc, str:
        The folder location of the EIS files (path to directory)
    area, float: 
        The active cell area in cm^2
    eis, bool: (default = True)
        If True, the EIS spectra are plotted
    drt, bool: (default = True)
        If True, the DRT spectra are plotted
    drt_peaks, bool: (default = True)
        If True, the DRT peaks are fit and plotted. If the peaks have not been fit, they
        will be fit and added the the cell data excel sheet. If the data exists in the spreadsheet already
        then that data will be plotted.
    ncol, int: (default = 1)
        The number of columns in the legend
    legend_loc, str: (default = 'best')
        The location of the legend. The other option is to put 
        the legend outside the figure and this is done by setting legend_loc = 'outside'
    overwrite, bool: (default = False)
        If the data already exists in the cell excel file, then this will overwrite that data with
    peaks_to_fit, str/int: (default: 'best_id')
        Sets the number of discrete elements in the EIS to fit in the DRT
        This basically just sets the amount of peaks to fit
        if this is set to the default 'best_id' then it fits the number of peaks suggested by dual_drt
        if this is set to a integer, it fit the number of peaks set by that integer
    drt_model, str: (default = dual)
        which type of drt used to analyze the data
        if drt = 'dual' the dual regression model is used to fit the data
        if drt = 'drtdop' then the drtdop model is used to fit the data
    save_eis, str: (default = None)
        If this is not none, the EIS figure will be saved.
        Save_eis is the file name and path of the saved file.
    save_drt, str: (default = None)
        If this is not none, , the DRT figure will be saved.
        Save_drt is the file name and path of the saved file.
        
    Return --> None but one or more plots are created and shown
    '''
    
    'Finding correct files and formatting'
    dta_files = [file for file in os.listdir(folder_loc) if file.endswith('.DTA')] #Makes a list of all .DTA files in the folder loc
    bias_eis = [] #initializing bias files list
    for file in dta_files: #Finding all fuel cell bias EIS files
        if (file.find('PEIS')!=-1) and (file.find('bias.DTA')!=-1) and (file.find('_n')!=-1):
            bias_eis.append(os.path.join(folder_loc,file))
    cell_name = os.path.basename(folder_loc).split('_')[2]
    bias_eis = sorted(bias_eis, key=lambda x: int((x[x.find('550C_n')+len('550C_n'):x.rfind('bias')]))) #Sorts numerically by bias

    for file in dta_files: #Finding the 0 bias condition and adding it to the beginning of the list of bias_eis
        if (file.find('PEIS')!=-1) and (file.find('0bias.DTA')!=-1): #adding in the 0 bias condition
            bias_eis.insert(0,os.path.join(folder_loc,file))

    'Plotting EIS'
    if eis == True:
        # --- Setting up the color map
        cmap = plt.cm.get_cmap('plasma') #cmr.redshift 
        color_space = np.linspace(0.2,0.8,len(bias_eis)) # array from 0-1 for the colormap for plotting
        c = 0 # index of the color array

        for peis in bias_eis:
            # --- Finding and formatting
            loc = os.path.join(folder_loc, peis) # creates the full path to the file
            bias = peis[peis.find('550C_')+len('PEIS_'):peis.rfind('bias')] #gets the bias from the file name
            if len(bias) > 1:
                bias = -int(bias[1:])/10

            nyquist_name =  str(bias) + 'V'

            # --- Plotting
            color = cmap(color_space[c])
            plot_peiss(area,nyquist_name,loc,ncol=ncol,legend_loc=legend_loc,color=color)
            c = c+1
        
        # - Saving the figure
        if save_eis is not None:
            fmat_eis = save_eis.split('.', 1)[-1]
            plt.savefig(save_eis, dpi=300, format=fmat_eis, bbox_inches='tight') 

        plt.show()

    ' ---------- inverting, plotting, and dual fitting all the FC bias DRT spectra'
    if drt == True: 
        fig, ax = plt.subplots() #initializing plots for DRT

        # --- Setting up the color map
        cmap = plt.cm.get_cmap('plasma') #cmr.redshift 
        color_space = np.linspace(0.2,0.8,len(bias_eis)) # array from 0-1 for the colormap for plotting
        c = 0 # index of the color array

        # --- Initializing lists for further analysis
        bias_array = np.array([]) # Bias of the cell (V) realtive to OCV
        ohmic_asr = np.array([]) # ohm*cm^2
        rp_asr = np.array([]) # ohm*cm^2
        df_tau_r = pd.DataFrame(columns = ['Bias (V)','Tau','Resistance']) #Initializing DataFrame to save temperature

        for peis in bias_eis: #For loop to plot the DRT of all PO2 files (probably a more elegant way, but this works)
            # --- Finding and formatting
            loc = os.path.join(folder_loc, peis) # creates the full path to the file
            bias = peis[peis.find('550C_')+len('PEIS_'):peis.rfind('bias')] #gets the bias from the file name
            if len(bias) > 1:
                bias = -int(bias[1:])/10
                
            label =  str(bias) + 'V'

            # --- Inverting the EIS data
            drt = DRT()
            df = read_eis(loc)
            freq,z = get_eis_tuple(df) # Get relavent data from EIS dataframe
            
            if drt_model == 'dual':
                drt.dual_fit_eis(freq, z, discrete_kw=dict(prior=True, prior_strength=None)) # Fit the data
                tau = drt.get_tau_eval(20)

                # - Selecting the number of peaks for the drt distribution to have.
                if peaks_to_fit == 'best_id':
                    best_id = drt.get_best_candidate_id('discrete', criterion='lml-bic')
                    peaks = best_id

                else: peaks = peaks_to_fit

                # - Selecting the model. The number of peaks to plot and fit
                model_dict = drt.get_candidate(peaks,'discrete')
                model = model_dict['model']

                # --- Plotting
                color = cmap(color_space[c])
                model.plot_distribution(tau, ax=ax, area=area,label=label,mark_peaks=True,color=color)
            
            elif drt_model == 'drtdop':
                drt.fit_dop=True
                drt.fit_eis(freq, z)
                tau = drt.get_tau_eval(20)
                
                # - Plotting
                color = cmap(color_space[c])

                mark_peaks_kw = {'color':color,'sizes':[75]}
                kwargs = {
                    'linewidth': 2,
                }
                drt.plot_distribution(tau, ax=ax, area=area,label=label,mark_peaks=True, c=color,
                                    scale_prefix="", plot_ci=False,mark_peaks_kw=mark_peaks_kw,
                                    **kwargs)
            else:
                print('In order to plot DRT, it must be fit with the dual or the drtdop model') 
                print('Set drt= \'dual\' or to \'drtdop\'') 
            
            c = c + 1 # Updating the color

            # --- Appending resistance values to lists
            bias_array = np.array([]) # Applied bias of the eis spectra relative to OCV
            ohmic_asr = np.append(ohmic_asr,drt.predict_r_inf() * area)
            rp_asr = np.append(rp_asr,drt.predict_r_p() * area)
            append_drt_peaks(df_tau_r, drt, area, bias, peaks_to_fit = peaks_to_fit,
                             drt_model = drt_model)

        # --- Formatting
        # - Adding Frequency scale:
        def Tau_to_Frequency(T):
            return 1 / (2 * np.pi * T)

        freq_ax = ax.secondary_xaxis('top', functions=(Tau_to_Frequency, Tau_to_Frequency))
        # - Excessive formatting
        ax_title_size = 23
        tick_label_size = ax_title_size * 0.8
        legend_txt_size = ax_title_size * 0.7

        ax.legend(fontsize=legend_txt_size,frameon=False,handletextpad=0.5,
                  handlelength=1)
        ax.xaxis.label.set_size(ax_title_size) 
        ax.yaxis.label.set_size(ax_title_size)
        ax.tick_params(axis='both', labelsize=tick_label_size)
        ax.spines['right'].set_visible(False)

        freq_ax.set_xlabel('$f$ (Hz)',size=ax_title_size)
        freq_ax.tick_params(axis='x',labelsize=tick_label_size)
    
        plt.tight_layout()

        if save_drt is not None:
            fmat_drt = save_drt.split('.', 1)[-1]
            fig.savefig(save_drt, dpi=300, format=fmat_drt, bbox_inches='tight') 

        plt.show()

        # >>>>>>>>>>>> Creating DataFrames and adding excel data sheets (or creating the file if need be)'
        # ------- Cell Resistance Data
        excel_name = '_' + cell_name + '_Data.xlsx'
        excel_file = os.path.join(folder_loc,excel_name)
        sheet_name = 'FC_bias_' + drt_model
        exists = False

        exists, writer = excel_datasheet_exists(excel_file,sheet_name)

        if exists == False:
            df_po2 = pd.DataFrame(list(zip(bias_array*100,ohmic_asr,rp_asr)),
                columns =['Bias (V)','Ohmic ASR (ohm*cm$^2$)', 'Rp ASR (ohm*cm$^2$)'])
            df_po2.to_excel(writer, sheet_name=sheet_name, index=False) # Writes this DataFrame to a specific worksheet
            writer.close() # Close the Pandas Excel writer and output the Excel file.

        elif exists == True and overwrite == True:
            df_po2 = pd.DataFrame(list(zip(bias_array*100,ohmic_asr,rp_asr)),
                columns =['Bias (V)','Ohmic ASR (ohm*cm$^2$)','Rp ASR (ohm*cm$^2$)'])
            
            book = load_workbook(excel_file)
            writer = pd.ExcelWriter(excel_file,engine='openpyxl',mode='a',if_sheet_exists='replace') #Creates a Pandas Excel writer using openpyxl as the engine in append mode
            df_po2.to_excel(writer, sheet_name=sheet_name, index=False) # Extract data to an excel sheet
            writer.close() # Close the Pandas Excel writer and output the Excel file.

            df_po2 = pd.read_excel(excel_file,sheet_name)

        # ------- Appending the Peak_fit data to an excel file
        excel_name = '_' + cell_name + '_Data.xlsx'
        excel_file = os.path.join(folder_loc,excel_name)
        peak_data_sheet = 'FC_bias_' + drt_model + '_DRT_peaks'
        exists_peaks = False

        exists_peaks, writer_peaks = excel_datasheet_exists(excel_file,peak_data_sheet)
    
        if exists_peaks == False: # Make the excel data list
            df_tau_r.to_excel(writer_peaks, sheet_name=peak_data_sheet, index=False) # Extract data to an excel sheet
            writer_peaks.close() # Close the Pandas Excel writer and output the Excel file.
        
        elif exists_peaks == True and overwrite == False: #load the data into a DataFrame
            df_tau_r = pd.read_excel(excel_file,peak_data_sheet)

        elif exists_peaks == True and overwrite == True:
            book = load_workbook(excel_file)
            writer_peaks = pd.ExcelWriter(excel_file,engine='openpyxl',mode='a',if_sheet_exists='replace') #Creates a Pandas Excel writer using openpyxl as the engine in append mode
            df_tau_r.to_excel(writer_peaks, sheet_name=peak_data_sheet, index=False) # Extract data to an excel sheet
            writer_peaks.close() # Close the Pandas Excel writer and output the Excel file.

            df_tau_r = pd.read_excel(excel_file,peak_data_sheet)

    ' --- DRT peak plotting --- '
    if drt_peaks == True and drt==True:
        # ----- plotting
        cmap = plt.cm.get_cmap('plasma_r') #cmr.redshift
        norm = Normalize(vmin=0.2, vmax=0.8)
        plot = sns.scatterplot(x = 'Tau', y = 'Resistance', data = df_tau_r, hue='Bias (V)',palette = cmap,norm=norm,s=69)

    elif drt_peaks == True:
        # --- Retrieving the data
        excel_name = '_' + cell_name + '_Data.xlsx'
        excel_file = os.path.join(folder_loc,excel_name)
        peak_data_sheet = 'FC_bias_dual_DRT_peaks'
        exists_peaks = False
        exists_peaks, writer_peaks = excel_datasheet_exists(excel_file,peak_data_sheet)
        
        if exists_peaks == True and overwrite == False: #load the data into a DataFrame
            df_tau_r = pd.read_excel(excel_file,peak_data_sheet)

        # --- Plotting
        cmap = plt.cm.get_cmap('plasma_r') #cmr.redshift
        cmap = ListedColormap(plt.get_cmap('plasma_r')(np.linspace(0.2, 0.8, 256))) 
        plot = sns.scatterplot(x = 'Tau', y = 'Resistance', data = df_tau_r, hue='Bias (V)',palette = cmap,s=69)

        # ----- Ascetics stuff
        sns.set_context("talk")
        fontsize = 14
        sns.despine()
        plot.set_ylabel('ASR (\u03A9 cm$^2$)',fontsize=fontsize)
        plot.set_xlabel('Time Constant (\u03C4/s)',fontsize=fontsize)
        plot.set(xscale='log')

        plt.tight_layout()
        plt.show()

    else:
        pass
    
def ec_bias_plots_dual(folder_loc:str, area:float, eis:bool=True,
                drt:bool=True, ncol:int=1, legend_loc:str='best',
                overwrite:bool = False, peaks_to_fit:int = 'best_id',
                drt_model:str = 'dual',save_eis:str=None,save_drt:str=None):
    '''
    Finds all EIS files in the folder_loc taken during bias testing in Electrolysis cell mode and plots the EIS
    The corresponding DRT fits are located and also plotted if drt = True
    All data is append to a sheet in the cell data excel file if it does not already exist. If the data
    does already exist that data is called upon for plotting

    The .DTA EIS files are taken during a Gamry sequence that I use for testing the cell performance under
    various biases

    Parameters:
    ------------
    folder_loc, str: 
        The folder location of the EIS files (path to directory)
    area, float: 
        The active cell area in cm^2
    eis, bool: 
        If True, the EIS spectra are plotted
    drt, bool: 
        If True, the DRT spectra are plotted
    ncol, int: 
        The number of columns in the legend
    legend_loc, str: 
        The location of the legend (default = 'best'). The other option is to put 
        the legend outside the figure and this is done by setting legend_loc = 'outside'
    overwrite, bool: (default = False)
        If the data already exists in the cell excel file, then this will overwrite that data with
    peaks_to_fit, str/int: (default: 'best_id')
        Sets the number of discrete elements in the EIS to fit in the DRT
        This basically just sets the amount of peaks to fit
        if this is set to the default 'best_id' then it fits the number of peaks suggested by dual_drt
        if this is set to a integer, it fit the number of peaks set by that integer
    drt_model, str: (default = dual)
        which type of drt used to analyze the data
        if drt = 'dual' the dual regression model is used to fit the data
        if drt = 'drtdop' then the drtdop model is used to fit the data
    save_eis, str: (default = None)
        If this is not none, the EIS figure will be saved.
        Save_eis is the file name and path of the saved file.
    save_drt, str: (default = None)
        If this is not none, , the DRT figure will be saved.
        Save_drt is the file name and path of the saved file.
        
    Return --> None but one or more plots are created and shown
    '''
    
    'Finding correct files and formatting'
    dta_files = [file for file in os.listdir(folder_loc) if file.endswith('.DTA')] #Makes a list of all .DTA files in the folder loc
    bias_eis = [] #initializing bias files list

    for file in dta_files: #Finding all fuel cell bias EIS files
        if (file.find('PEIS')!=-1) and (file.find('bias.DTA')!=-1) and (file.find('_n')==-1):
            bias_eis.append(os.path.join(folder_loc,file))

    cell_name = os.path.basename(folder_loc).split('_')[2]
    bias_eis = sorted(bias_eis, key=lambda x: int((x[x.find('550C_')+len('550C_'):x.rfind('bias')]))) #Sorts numerically by bias

    'Plotting EIS'
    if eis == True:
        for peis in bias_eis:
            # --- Finding and formatting
            loc = os.path.join(folder_loc, peis) # creates the full path to the file
            bias = peis[peis.find('550C_')+len('PEIS_'):peis.rfind('bias')] #gets the bias from the file name
            bias = int(bias)/10

            nyquist_name =  str(bias) + 'V'

            # --- Plotting
            plot_peiss(area,nyquist_name,loc,ncol=ncol,legend_loc=legend_loc)

        # - Saving the figure
        if save_eis is not None:
            fmat_eis = save_eis.split('.', 1)[-1]
            plt.savefig(save_eis, dpi=300, format=fmat_eis, bbox_inches='tight')  
        
        plt.show()

    'Plotting DRT'
    if drt == True:
        fig, ax = plt.subplots() #initializing plots for DRT

        # --- Initializing lists of data to save
        bias_array = np.array([]) # Applied bias of the eis spectra relative to OCV
        ohmic_asr = np.array([]) # ohm*cm^2
        rp_asr = np.array([]) # ohm*cm^2

        for peis in bias_eis: # For loop to plot the DRT of all PO2 files (probably a more elegant way, but this works)
            # --- Finding and formatting
            loc = os.path.join(folder_loc, peis) # creates the full path to the file
            bias = peis[peis.find('550C_')+len('PEIS_'):peis.rfind('bias')] #gets the bias from the file name
            map_fit_name = cell_name + '_map_fit_' + bias + 'bias.pkl'
            if int(bias) > 0:
                bias = int(bias)/10

            label =  str(bias) + 'V'
            
            # --- Inverting the EIS data
            df = read_eis(loc)
            freq,z = get_eis_tuple(df) # Get relavent data from EIS dataframe
            drt = DRT()
            
            # --- Plotting the DRT
            if drt_model == 'dual':
                drt.dual_fit_eis(freq, z, discrete_kw=dict(prior=True, prior_strength=None), nonneg=False) # Fit the data
                
                # drt.plot_distribution(mark_peaks=True, label=label, ax=ax, area=area)
                tau = drt.get_tau_eval(20)

                # - Selecting the number of peaks for the drt distribution to have.
                if peaks_to_fit == 'best_id':
                    best_id = drt.get_best_candidate_id('discrete', criterion='lml-bic')
                    peaks = best_id

                else: peaks = peaks_to_fit

                # - Selecting the model. The number of peaks to plot and fit
                model_dict = drt.get_candidate(peaks,'discrete')
                model = model_dict['model']

                # --- Plotting
                model.plot_distribution(tau, ax=ax, area=area,label=label,mark_peaks=True)
            
            elif drt_model == 'drtdop':
                # drt = plot_drtdop(loc, area, label=label, ax=ax, mark_peaks=True, nonneg=False, scale_prefix = "")
                drt.fit_dop=True
                drt.fit_eis(freq, z, nonneg=False)
                tau = drt.get_tau_eval(20)
                
                # - Plotting
                # color = cmap(color_space[c])

                mark_peaks_kw = {'sizes':[75]}
                kwargs = {
                    'linewidth': 2,
                }
                drt.plot_distribution(tau, ax=ax, area=area,label=label,mark_peaks=True,
                                    scale_prefix="", plot_ci=False,mark_peaks_kw=mark_peaks_kw,
                                    **kwargs)
            else:
                print('In order to plot DRT, it must be fit with the dual or the drtdop model') 
                print('Set drt= \'dual\' or to \'drtdop\'')          

            # --- Appending resistance values to lists
            bias_array = np.array([]) # Applied bias of the eis spectra relative to OCV
            ohmic_asr = np.append(ohmic_asr,drt.predict_r_inf() * area)
            rp_asr = np.append(rp_asr,drt.predict_r_p() * area)

        # --- Formatting
        # - Adding Frequency scale:
        def Tau_to_Frequency(T):
            return 1 / (2 * np.pi * T)

        freq_ax = ax.secondary_xaxis('top', functions=(Tau_to_Frequency, Tau_to_Frequency))
        
        # - Excessive formatting
        ax_title_size = 21
        tick_label_size = ax_title_size * 0.8
        legend_txt_size = ax_title_size * 0.7

        ax.legend(fontsize=legend_txt_size,frameon=False,handletextpad=0.5,
                  handlelength=1,loc='center left',bbox_to_anchor=(0.95,0.5))
        ax.xaxis.label.set_size(ax_title_size) 
        ax.yaxis.label.set_size(ax_title_size)
        ax.tick_params(axis='both', labelsize=tick_label_size)
        ax.spines['right'].set_visible(False)

        freq_ax.set_xlabel('$f$ (Hz)',size=ax_title_size)
        freq_ax.tick_params(axis='x',labelsize=tick_label_size)
        
        plt.tight_layout()

        if save_drt is not None:
            fmat_drt = save_drt.split('.', 1)[-1]
            fig.savefig(save_drt, dpi=300, format=fmat_drt, bbox_inches='tight') 

        plt.show()


        # >>>>>>>>>>>> Creating DataFrames and adding excel data sheets (or creating the file if need be)'
        # ------- Cell Resistance Data
        excel_name = '_' + cell_name + '_Data.xlsx'
        excel_file = os.path.join(folder_loc,excel_name)
        sheet_name = 'EC_bias_' + drt_model
        exists = False

        exists, writer = excel_datasheet_exists(excel_file,sheet_name)

        if exists == False:
            df_po2 = pd.DataFrame(list(zip(bias_array*100,ohmic_asr,rp_asr)),
                columns =['Bias (V)','Ohmic ASR (ohm*cm$^2$)', 'Rp ASR (ohm*cm$^2$)'])
            df_po2.to_excel(writer, sheet_name=sheet_name, index=False) # Writes this DataFrame to a specific worksheet
            writer.close() # Close the Pandas Excel writer and output the Excel file.

        elif exists == True and overwrite == True:
            df_po2 = pd.DataFrame(list(zip(bias_array*100,ohmic_asr,rp_asr)),
                columns =['Bias (V)','Ohmic ASR (ohm*cm$^2$)','Rp ASR (ohm*cm$^2$)'])
            
            book = load_workbook(excel_file)
            writer = pd.ExcelWriter(excel_file,engine='openpyxl',mode='a',if_sheet_exists='replace') #Creates a Pandas Excel writer using openpyxl as the engine in append mode
            df_po2.to_excel(writer, sheet_name=sheet_name, index=False) # Extract data to an excel sheet
            writer.close() # Close the Pandas Excel writer and output the Excel file.

            df_po2 = pd.read_excel(excel_file,sheet_name)

def o2_dual_drt_peaks(folder_loc:str, tau_low:float, tau_high:float, concs:np.array = None,
                        rmv_concs_r:np.array = None, rmv_concs_l:np.array = None,plot_all=False,
                        drt_model:str = 'dual',print_c:bool = True):
    '''
    This function is meant to linearly fit a single DRT peak across an oxygen concentration range
    This function can only be used after the po2_plots function
    This function plots ln(1/asr) vs. ln(O2 concentration (%))

    Parameters
    ----------
    folder_loc, str: (path to a directory)
        The location of the directory containing the EIS files.
    tau_low, float:
        The lower bound of the time constant range to fit.
    tau_high, float:
        The upper bound of the time constant range to fit.
    concs, np.array: (default = None)
        The concentrations that the DRT peaks are taken at. if this is none all temperature ranges in
        the cell data sheet will be fit.
    rmv_concs_r, np.array: (default = None)
        If two clusters overlap, specify the concentrations where there are overlap and this will remove
        the peaks with higher time constants (lower frequency, to the right) from the fit.
    rmv_concs_l, np.array: (default = None)
        If two clusters overlap, specify the concentrations where there are overlap and this will remove
        the peaks with lower time constants (higher frequency, to the left) from the fit.
    drt_model, str: (default = dual)
        which type of drt used to analyze the data
        if drt = 'dual' the dual regression model is used to fit the data
        if drt = 'drtdop' then the drtdop model is used to fit the data
    print_c , bool: (default = False)
        if desired, the capacitance of each peak is calculated and fit

    Returns --> The slope of ln(1/asr)/ln(O2%), and a plot of the DRT peak fit and the activation energy is calculated and printed on the plot
    '''

    # ----- Sorting out the points from a specific time constant
    for file in os.listdir(folder_loc):
        if file.endswith('Data.xlsx'):
            data_file = os.path.join(folder_loc,file)
            break
    
    sheet_name = 'pO2_' + drt_model + '_DRT_peaks'
    all_data = pd.read_excel(data_file,sheet_name)
    data = all_data[(all_data['Tau']>tau_low) & (all_data['Tau']<tau_high)]

    # ----- If desired, plotting all of the PO2 fits
    if plot_all == True:
        # ----- plotting
        palette = sns.color_palette('crest_r', as_cmap=True)
        plot = sns.scatterplot(x = 'Tau', y = 'Resistance', data = all_data, hue='O2 Concentration (%)',palette = palette,s=69)

        # ----- Ascetics stuff
        sns.set_context("talk")
        fontsize = 14
        sns.despine()
        plot.set_ylabel('ASR (\u03A9 cm$^2$)',fontsize=fontsize)
        plot.set_xlabel('Time Constant (\u03C4/s)',fontsize=fontsize)
        plot.set(xscale='log')

        plt.tight_layout()
        plt.show()

    # ----- If desired, only plotting certain temperatures
    if concs is not None:
        data = data[data['O2 Concentration (%)'].isin(concs)]

    # ----- removing a duplicate if there is overlap between clusters to the right (remove points to the right)
    if rmv_concs_r is not None:
        for conc in rmv_concs_r:
            # - find the lowest tau for a given temperature
            conc_data = data[data['O2 Concentration (%)']==conc]
            conc_data = conc_data.sort_values(by='Tau')

            try:
                highest_tau = conc_data.iloc[-1]['Tau']

            except IndexError as error:
                traceback.print_exc()
                print('The removed concentration must be in concs array (aka concentrations plotted)')
                print('Check the concs array and the rmv_concs array')
                print('If that does not work, make sure that there are data points in the specified tau range (it is easy to confuse the log scale)')
                sys.exit(1)

            # - remove all higher tau values
            # data = data[data['Tau']!=highest_tau]
            data = data[~((data['O2 Concentration (%)'] == conc) & (data['Tau'] == highest_tau))]


    # ----- removing a duplicate if there is overlap between clusters to the left (remove points to the left)
    if rmv_concs_l is not None:
        for conc in rmv_concs_l:
            # - find the lowest tau for a given temperature
            conc_data = data[data['O2 Concentration (%)']==conc]
            conc_data = conc_data.sort_values(by='Tau')

            try:
                lowest_tau = conc_data.iloc[0]['Tau']

            except IndexError as error:
                traceback.print_exc()
                print('The removed concentration must be in concs array (aka concentrations plotted)')
                print('Check the concs array and the rmv_concs array')
                print('If that does not work, make sure that there are data points in the specified tau range (it is easy to confuse the log scale)')
                sys.exit(1)

            # - remove all higher tau values
            # data = data[data['Tau']!=lowest_tau]
            data = data[~((data['O2 Concentration (%)'] == conc) & (data['Tau'] == lowest_tau))]

    
    # ----- Finding low and high Tau values of the values plotted
    min_tau = data['Tau'].min()
    max_tau = data['Tau'].max()

    # ----- Formatting data for plotting
    rp_asr = data['Resistance'].values
    concs = data['O2 Concentration (%)'].values
    ln_rp_asr = np.log(1/rp_asr)
    
    o2_concs = np.array(concs)
    ln_o2 = np.log(o2_concs/100)

    # --- If printing capacitance is desired
    data['Capacitance (F/cm^2)'] = data['Tau']/data['Resistance']
    if print_c == True:
        print(data)

    # ----- Plotting
    x = ln_o2
    y = ln_rp_asr

    fig, ax1 = plt.subplots()
    ax2 = ax1.twiny()
    ax1.plot(x,y,'ko')

    # - Setting font sizes
    label_fontsize = 'x-large'
    tick_fontsize = 'large'
    text_fontsize = 'x-large'

    # - Setting ticks
    ax1.set_xticks(x, np.round(x,2), fontsize=tick_fontsize)
    ax1.set_yticks(y,labels= np.round(y,2), fontsize=tick_fontsize)
    ax2.set_xticks(x, labels = concs, fontsize=tick_fontsize)
    ax1.set_xlim(-1.7,0.1)
    ax2.set_xlim(-1.7,0.1)

    # - linear Fit:
    mr,br = np.polyfit(ln_o2,ln_rp_asr,1)
    mr,br, r, p_value, std_err = scipy.stats.linregress(x, y)
    fit_r = mr*ln_o2 + br
    ax1.plot(x, fit_r,'g')

    # - Axis labels:
    ax1.set_xlabel('ln(O$_2$) (%)',fontsize=label_fontsize)
    ax2.set_xlabel('Oxygen Concentration (%)',fontsize=label_fontsize)
    ax1.set_ylabel('ln(1/ASR) (S/cm$^2$)',fontsize=label_fontsize)

    # - Creating table:
    row_labels = ['r squared']
    slopes = f'{round(mr,2)}'
    table_values = [[round(r**2,3)]] # $\\bf{round(mr,2)}$
    color = 'palegreen'

    if mr >= 0:
        table = plt.table(cellText=table_values,colWidths = [.2]*3,rowLabels=row_labels,loc = 'lower right',
            rowColours= [color,color])
    else:
        table = plt.table(cellText=table_values,colWidths = [.2]*3,rowLabels=row_labels,loc = 'lower center',
            rowColours= [color,color])

    table.scale(1,1.6)
    
    # - Printing figure text
    mr_str = f'{round(mr,2)}'
    # - Time constants
    tau_lows = f'{min_tau:.2e}'
    tau_highs = f'{max_tau:.2e}'
    # - Capacitances
    avg_c = data['Capacitance (F/cm^2)'].mean()
    std_c = data['Capacitance (F/cm^2)'].std()
    avg_cs = f'{avg_c:.2e}'
    std_cs = f'{std_c:.2e}'

    if mr >=0:
        fig.text(0.68,0.22,r'ASR$_\mathrm{P}$ Slope = '+mr_str,weight='bold',fontsize = tick_fontsize)
        fig.text(0.17,0.81,'DRT peak between '+tau_lows+'(\u03C4/s) and '+tau_highs+'(\u03C4/s)',fontsize = tick_fontsize)
        fig.text(0.17,0.76,'Avg C: '+avg_cs+r'(F/cm$^2$) +/- '+std_cs,fontsize = tick_fontsize)

    else:
        fig.text(0.37,0.22,r'ASR$_\mathrm{P}$ Slope = '+mr_str,weight='bold',fontsize = tick_fontsize)
        fig.text(0.38,0.81,'DRT peak between '+tau_lows+'(\u03C4/s) and '+tau_highs+'(\u03C4/s)',fontsize = tick_fontsize)
        fig.text(0.38,0.76,'Avg C: '+avg_cs+r'(F/cm$^2$) +/- '+std_cs,fontsize = tick_fontsize)


    plt.tight_layout()

    plt.show()

    return(mr)

def o2_pfrt_drt_peaks(folder_loc:str, tau_low:float, tau_high:float, concs:np.array = None,
                        rmv_concs_r:np.array = None, rmv_concs_l:np.array = None):
    '''
    This function is meant to linearly fit a single DRT peak across an oxygen concentration range
    This function can only be used after the po2_plots function
    This function plots ln(1/asr) vs. ln(O2 concentration (%))

    Parameters
    ----------
    folder_loc, str: (path to a directory)
        The location of the directory containing the EIS files.
    tau_low, float:
        The lower bound of the time constant range to fit.
    tau_high, float:
        The upper bound of the time constant range to fit.
    concs, np.array: (default = None)
        The concentrations that the DRT peaks are taken at. if this is none all temperature ranges in
        the cell data sheet will be fit.
    rmv_concs_r, np.array: (default = None)
        If two clusters overlap, specify the concentrations where there are overlap and this will remove
        the peaks with higher time constants (lower frequency, to the right) from the fit.
    rmv_concs_l, np.array: (default = None)
        If two clusters overlap, specify the concentrations where there are overlap and this will remove
        the peaks with lower time constants (higher frequency, to the left) from the fit.

    Returns --> The slope of ln(1/asr)/ln(O2%), and a plot of the DRT peak fit and the activation energy is calculated and printed on the plot
    '''

    # ----- Sorting out the points from a specific time constant
    for file in os.listdir(folder_loc):
        if file.endswith('Data.xlsx'):
            data_file = os.path.join(folder_loc,file)
            break
    
    data = pd.read_excel(data_file,'pO2_PFRT_DRT_peak_fits')
    data = data[(data['Tau']>tau_low) & (data['Tau']<tau_high)]

    # ----- If desired, only plotting certain temperatures
    if concs is not None:
        data = data[data['O2 Concentration (%)'].isin(concs)]

    # ----- removing a duplicate if there is overlap between clusters to the right (remove points to the right)
    if rmv_concs_r is not None:
        for conc in rmv_concs_r:
            # - find the lowest tau for a given temperature
            conc_data = data[data['O2 Concentration (%)']==conc]
            conc_data = conc_data.sort_values(by='Tau')

            try:
                highest_tau = conc_data.iloc[-1]['Tau']

            except IndexError as error:
                traceback.print_exc()
                print('The removed concentration must be in concs array (aka concentrations plotted)')
                print('Check the concs array and the rmv_concs array')
                print('If that does not work, make sure that there are data points in the specified tau range (it is easy to confuse the log scale)')
                sys.exit(1)

            # - remove all higher tau values
            data = data[data['Tau']!=highest_tau]

    # ----- removing a duplicate if there is overlap between clusters to the left (remove points to the left)
    if rmv_concs_l is not None:
        for conc in rmv_concs_l:
            # - find the lowest tau for a given temperature
            conc_data = data[data['O2 Concentration (%)']==conc]
            conc_data = conc_data.sort_values(by='Tau')

            try:
                lowest_tau = conc_data.iloc[0]['Tau']

            except IndexError as error:
                traceback.print_exc()
                print('The removed concentration must be in concs array (aka concentrations plotted)')
                print('Check the concs array and the rmv_concs array')
                print('If that does not work, make sure that there are data points in the specified tau range (it is easy to confuse the log scale)')
                sys.exit(1)

            # - remove all higher tau values
            data = data[data['Tau']!=lowest_tau]
    
    # ----- Finding low and high Tau values of the values plotted
    min_tau = data['Tau'].min()
    max_tau = data['Tau'].max()

    # ----- Formatting data for plotting
    rp_asr = data['Resistance'].values
    concs = data['O2 Concentration (%)'].values
    ln_rp_asr = np.log(1/rp_asr)
    
    o2_concs = np.array(concs)
    ln_o2 = np.log(o2_concs/100)

    # ----- Plotting
    x = ln_o2
    y = ln_rp_asr

    fig, ax1 = plt.subplots()
    ax2 = ax1.twiny()
    ax1.plot(x,y,'ko')

    # - Setting font sizes
    label_fontsize = 'x-large'
    tick_fontsize = 'large'
    text_fontsize = 'x-large'

    # - Setting ticks
    ax1.set_xticks(x, np.round(x,2), fontsize=tick_fontsize)
    ax1.set_yticks(y,labels= np.round(y,2), fontsize=tick_fontsize)
    ax2.set_xticks(x, labels = concs, fontsize=tick_fontsize)
    ax1.set_xlim(-1.7,0.1)
    ax2.set_xlim(-1.7,0.1)

    # - linear Fit:
    mr,br = np.polyfit(ln_o2,ln_rp_asr,1)
    mr,br, r, p_value, std_err = scipy.stats.linregress(x, y)
    fit_r = mr*ln_o2 + br
    ax1.plot(x, fit_r,'g')

    # - Axis labels:
    ax1.set_xlabel('ln(O$_2$) (%)',fontsize=label_fontsize)
    ax2.set_xlabel('Oxygen Concentration (%)',fontsize=label_fontsize)
    ax1.set_ylabel('ln(1/ASR) (S/cm$^2$)',fontsize=label_fontsize)

    # - Creating table:
    row_labels = ['r squared']
    slopes = f'{round(mr,2)}'
    table_values = [[round(r**2,3)]] # $\\bf{round(mr,2)}$
    color = 'palegreen'

    if mr >= 0:
        table = plt.table(cellText=table_values,colWidths = [.2]*3,rowLabels=row_labels,loc = 'lower right',
            rowColours= [color,color])
    else:
        table = plt.table(cellText=table_values,colWidths = [.2]*3,rowLabels=row_labels,loc = 'lower center',
            rowColours= [color,color])

    table.scale(1,1.6)
    
    # - Printing figure text
    mr_str = f'{round(mr,2)}'
    tau_lows = f'{min_tau:.2e}'
    tau_highs = f'{max_tau:.2e}'
    if mr >=0:
        fig.text(0.68,0.22,r'ASR$_\mathrm{P}$ Slope = '+mr_str,weight='bold',fontsize = tick_fontsize)
        fig.text(0.17,0.81,'DRT peak between '+tau_lows+'(\u03C4/s) and '+tau_highs+'(\u03C4/s)',fontsize = tick_fontsize)
    else:
        fig.text(0.37,0.22,r'ASR$_\mathrm{P}$ Slope = '+mr_str,weight='bold',fontsize = tick_fontsize)
        fig.text(0.38,0.81,'DRT peak between '+tau_lows+'(\u03C4/s) and '+tau_highs+'(\u03C4/s)',fontsize = tick_fontsize)
    plt.tight_layout()

    plt.show()

    return(mr)

def excel_datasheet_exists(excel_file:str,sheet_name:str):
    '''
    Finds if the excel file exists and if the sheet name exists in the excel file
    If the excel file does not exist, it will return a writer that can create excel files
    If the excel file does exist, it will return a writer that can append to the excel file
    Will return whether or not the sheet_name exists in the excel file

    Parameters:
    ------------
    excel_file, str:
        The excel file to look for
    sheet_name, str: 
        The name of the excel sheet to look for

    Return --> appropriate writer, whether or not the excel sheet exists
    '''
    
    exists = False
    if os.path.exists(excel_file)==True: # Looking for the data excel file in the button cell folder
        writer = pd.ExcelWriter(excel_file,engine='openpyxl',mode='a') #Creates a Pandas Excel writer using openpyxl as the engine in append mode
        wb = load_workbook(excel_file, read_only=True) # Looking for the bias Deg eis
        
        if sheet_name in wb.sheetnames:
            exists = True

    elif os.path.exists(excel_file)==False:
        writer = pd.ExcelWriter(excel_file,engine='xlsxwriter')
    
    return exists,writer

def append_drt_peaks(df_tau_r, drt, area, condition, peaks_to_fit = 'best_id',drt_model='dual'):
    '''
    Takes a DRT instance and extracts the peak data from it. It then appends this data to a dataframe (df_tau_r)
    The dataframe is made outside the function
    
    Parameters:
    ----------
    df_tau_r, pandas dataframe:
        dataframe to append the drt peak tau and resistance data to
    drt, drt instance:
        DRT instance
    area, float:
        The active cell area in cm^2
    condition, float:
        the condition of interest that the drt data was taken in
    peaks_to_fit, str/int: (default: 'best_id')
        Sets the number of discrete elements in the EIS to fit in the DRT
        This basically just sets the amout of peaks to fit
        if this is set to the default 'best_id' then it fits the number of peaks suggested by dual_drt
        if this is set to a integer, it fit the number of peaks set by that integer
    drt_model, str: (default = dual)
        which type of drt used to analyze the data
        if drt = 'dual' the dual regression model is used to fit the data
        if drt = 'drtdop' then the drtdop model is used to fit the data
    
    Return --> None
    '''
    if drt_model == 'dual':
        # --- obtain time constants from inverters and Appending tau and r for each peak into df_tau_r
        if peaks_to_fit == 'best_id':
            best_id = drt.get_best_candidate_id('discrete', criterion='lml-bic')
            peaks = best_id

        else: peaks = peaks_to_fit
        # - Selecting the model. The number of peaks to plot and fit

        model_dict = drt.get_candidate(peaks,'discrete')
        model = model_dict['model']
        tau = model_dict['peak_tau'] # τ/s

        # - Obtaining the resistance value for each peak
        r_list = []
        for i in range(1,int(peaks)+1):
            peak = 'R_HN'+str(i)
            resistance = model.parameter_dict[peak]
            r_list.append(resistance)

    elif drt_model == 'drtdop':
        tau = drt.find_peaks().tolist() # Finding Tau, and making it iterable
        r_list = drt.quantify_peaks(tau=tau)

    r = np.array(r_list) * area # Ω*cm2

    i = 0
    for τ in tau:
        df_tau_r.loc[len(df_tau_r.index)] = [condition, τ, r[i]]
        i = i+1

def rp_ohmic_to_excel(cell_name, folder_loc, ohmic_asr, rp_asr, sheet_name, condition, condition_label, overwrite = False):
    '''
    Takes in ohmic and rp data calculated from EIS spectra and writes it to an excel file
    The data comes in as lists, is zipped to a dataframe, and is written to an excel file

    Parameters:
    ----------
    cell_name, str:
        Name of the cell that the data was taken from
    folder_loc, str:
        The folder location of the EIS files (path to directory)
    ohmic_asr, list:
        list of the ohmic area specific resistance (asr) values for the cell
        list of floats
    rp_asr, list:
        list of the polarization (Rp) area specific resistance (asr) values for the cell
        list of floats
    sheet_name, str: 
        The name for the excel sheet that is about to be created
    condition, list:
        list of the condition that is varying during testing
        could be time, bias, PO2, etc...
        the condition values will match up with the Rp and ohmic values
        list of floats
    condition_label, str:
        The name of the column in the dataframe containing the condition data
    overwrite, bool: (default = False)
        If the data already exists in the cell excel file, then this will overwrite that data with
        the data currently being fit when overwrite=True.

    Return --> None, but the data is written to an excel file
    '''
    
    # >>>>>>>>>>>> Creating DataFrames and adding excel data sheets (or creating the file if need be)'
    excel_name = '_' + cell_name + '_Data.xlsx'
    excel_file = os.path.join(folder_loc,excel_name)
    exists = False

    exists, writer = excel_datasheet_exists(excel_file,sheet_name)

    if exists == False:
        df = pd.DataFrame(list(zip(condition,ohmic_asr,rp_asr)),
            columns =[condition_label,'Ohmic ASR (ohm*cm$^2$)', 'Rp ASR (ohm*cm$^2$)'])
        df.to_excel(writer, sheet_name=sheet_name, index=False) # Writes this DataFrame to a specific worksheet
        writer.close() # Close the Pandas Excel writer and output the Excel file.

    elif exists == True and overwrite == True:
        df = pd.DataFrame(list(zip(condition,ohmic_asr,rp_asr)),
            columns =[condition_label,'Ohmic ASR (ohm*cm$^2$)', 'Rp ASR (ohm*cm$^2$)'])
        
        book = load_workbook(excel_file)
        writer = pd.ExcelWriter(excel_file,engine='openpyxl',mode='a',if_sheet_exists='replace') #Creates a Pandas Excel writer using openpyxl as the engine in append mode
        df.to_excel(writer, sheet_name=sheet_name, index=False) # Extract data to an excel sheet
        writer.close() # Close the Pandas Excel writer and output the Excel file.

def df_tau_r_to_excel(cell_name, folder_loc, df_tau_r,sheet_name, overwrite = False):
    '''
    Takes in the DRT peak fit data and writes it to an excel file
    The data comes in as a dataframe and is written to an excel file

    Parameters:
    ----------
    cell_name, str:
        Name of the cell that the data was taken from
    folder_loc, str:
        The folder location of the EIS files (path to directory)
    df_tau_r, pandas df:
        dataframe containing the tau, and resistance data for each DRT peak for each condition
    sheet_name, str: 
        The name for the excel sheet that is about to be created
    overwrite, bool: (default = False)
        If the data already exists in the cell excel file, then this will overwrite that data with
        the data currently being fit when overwrite=True.

    Return --> None, but the data is written to an excel file
    '''

    # ------- Appending the Peak_fit data to an excel file
    excel_name = '_' + cell_name + '_Data.xlsx'
    excel_file = os.path.join(folder_loc,excel_name)
    exists_peaks = False

    exists_peaks, writer_peaks = excel_datasheet_exists(excel_file,sheet_name)

    if exists_peaks == False: # Make the excel data list
        df_tau_r.to_excel(writer_peaks, sheet_name=sheet_name, index=False) # Extract data to an excel sheet
        writer_peaks.close() # Close the Pandas Excel writer and output the Excel file.

    elif exists_peaks == True and overwrite == True:
        book = load_workbook(excel_file)
        writer_peaks = pd.ExcelWriter(excel_file,engine='openpyxl',mode='a',if_sheet_exists='replace') #Creates a Pandas Excel writer using openpyxl as the engine in append mode
        df_tau_r.to_excel(writer_peaks, sheet_name=sheet_name, index=False) # Extract data to an excel sheet
        writer_peaks.close() # Close the Pandas Excel writer and output the Excel file.

def plot_drt_peaks(cell_folder:str, sheet_name:str, condition_label:str, ax = None, cmap = 'viridis',test_type = None):
    '''
    Plot drt peaks for all DRT data taken from a cell over a certain condition
    The data is pulled from the cell data spread sheet
    Thus, the DRT needs to already be fit, analyzed, and appended to the datasheet

    Parameters:
    ----------
    cell_folder, str:
        Path to the folder that contains the data and readme for each cell
    sheet_name, str:
        Name of the datasheet in the cell data excel file that contains the desired data
    ax, mpl.Axes: (default = None)
        Axes object used to plot the figure
    condition_label, str:
        The name of the column in the dataframe containing the condition data
    test_type, str: (default = None)
        The type of test ran to get the data
        ex: 'stb', ...
    cmap, str or cmap object: (default = 'viridis')
        The colormap used for plotting

    Return --> None, but a figure is plot
    '''
    # --- Taking the correct data from the cell Data spreadsheet
    cell_name = find_cell_name(cell_folder)
    excel_name = '_' + cell_name + '_Data.xlsx'
    excel_file = os.path.join(cell_folder,excel_name)
    df = pd.read_excel(excel_file, sheet_name)

    # ----- plotting
    if ax is None: # Idea for this If statement taken from Dr. Jake Huang's code
        fig, ax = plt.subplots()
    else:
        fig = ax.get_figure()
        
    cmap = plt.cm.get_cmap(cmap)
    ax = sns.scatterplot(x = 'Tau', y = 'Resistance', data = df, hue= condition_label ,palette = cmap,s=69, legend=False)

    # ---- Formatting
    ax.set_ylabel('Peak ASR (\u03A9 cm$^2$)',fontsize='xx-large')
    ax.set_xlabel('Time Constant (\u03C4/s)',fontsize='xx-large')
    ax.set(xscale='log')
    ax.set_ylim(0)

    # --- Excessive formatting
    sns.despine()
    ax.tick_params(axis='both', which='major', labelsize='x-large')
    ax.yaxis.set_major_locator(ticker.MaxNLocator(nbins=5))
    ax.yaxis.set_major_formatter(ticker.ScalarFormatter())

    if test_type == 'stb':
        end_time = df[condition_label].iloc[-1]
        # --- Colorbar formatting
        sm = plt.cm.ScalarMappable(cmap=cmap)
        sm.set_array([0,end_time])
        divider = make_axes_locatable(ax)
        cax = divider.append_axes("right", size="5%", pad=0.05)
        cb = plt.colorbar(sm,ticks = [0,end_time],cax=cax)
        cb.set_label(label='Time (hrs)',fontsize = 'xx-large',labelpad = -10)
        # cb.set_ticks(ticks = [0,end_time],fontsize='xx-large')
        cb.ax.tick_params(labelsize='x-large')

    plt.tight_layout()

def find_cell_name(folder_loc):
    """
    Finds the name of the cell

    Parameters
    ----------
    folder_loc, str: (path to a directory)
        The folder location of the EIS files 

    Return --> str, The name of the cell
    """
    for file in os.listdir(folder_loc): # Finding the name of the cell
        if file.find('PEIS')!=-1:
            cell_name = os.path.basename(file).split("_", 1)[0]
            break
        
    return(cell_name)

def clamp_resistance(galv_loc):
    '''
    Returns and prints the clamp resistance value
    Measured using a galvanostatic hold with a Gamry potentiostat
    Uses Ohms law (V=IR)

    Parameters:
    ----------
    galv_loc, str:
        path to the galvanostatic .dta file
        
    Return --> R, float
    '''
    galv_txt = read_dta(galv_loc)

    # --- Finding the start of the data (Also informed by Hybrid_drt)
    data_flag = galv_txt.find('COMPLIANCEVOLTAGE')
    metadata = galv_txt[:data_flag] # figure out where the metadata ends
    skiprows = len(metadata.split('\n')) + 2 # Finding how many rows to skip before the data starts

    df = pd.read_csv(galv_loc,sep='\t',skiprows=skiprows,encoding='latin1',engine='python')

    # --- Calculations
    V = df['V vs. Ref.'].mean(axis=0) # V
    I = df['A'].mean(axis=0) # A
    R = V/I

    # Printing and returning the resistance value
    print(f"The resistance of the clamp is: {R:.4f}")

    return R






