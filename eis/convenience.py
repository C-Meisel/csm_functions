''' This module contains functions to format, plot, analyze, and save EIS and DRT spectra for multiple 
Electrochemical cell tests.  The data comes from a Gamry Potentiostat
# C-Meisel
'''
# import bayes_drt2
'Imports'
import os
import pandas as pd
import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
import natsort
import seaborn as sns
import scipy as scipy
import cmasher as cmr
import sys # Used to go through the system
import traceback # Used to look at errors
from mpl_toolkits.axes_grid1.axes_divider import make_axes_locatable
from bayes_drt2.inversion import Inverter#inverter class in inversion module
from bayes_drt2 import file_load as fl
from bayes_drt2 import plotting as bp #Imports new plotting module (bp = bayes plotting)
from mpl_toolkits.axes_grid1.axes_divider import make_axes_locatable


from .plotting import plot_peis, plot_peiss, lnpo2
from .fit_drt import map_drt_save
from .data_formatting import peis_data


def plot_resistanceDeg(folder_loc:str, jar_bias:str, jar_ocv:str, area:float, 
                        plot_ocv:bool = True, title:bool = True): #Recently has not been working Jan21
    '''
    Plots the ohmic, polarization (Rp), and total resistance of a cell over time.
    The data is obtained from a gamry sequence that periodically takes EIS of a cell over time
    The data must then be fit by the function "deg_eis_fitter" before using this function.
    If eis is taken at OCV and Bias, this function will make a plot for each condition.

    This function also saves the resistance data to a sheet in an excel file for this cell
    If the excel file does not exist, this function will create it. 
    If the sheet already exists, this function will plot the existing data in the sheet.

    param folder_loc, string: 
        location of the folder containing the EIS stability data
    param jar_bias, string: 
        path to the jar to save the bias DRT fits to
    param jar_ocv, string: 
        path to the jar to save the ocv DRT fits to
    param area, string: 
        The active cell area in cm^2
    param plot_ocv, boolean: (default = True)
        Whether to plot the OCV data
    param title, boolean: (default = True)
        Whether to give the plots titles

    Return --> None
    '''
    
    files = os.listdir(folder_loc)
    cell_name = os.path.basename(folder_loc).split("_", 3)[2]
    peis_deg_files = [file for file in files if file.endswith('.DTA') and file.find('Deg')!=-1 and file.find('PEIS')!=-1] #Makes a list of all the degradation EIS files
    
    'Extracting and plotting bias results'
    # ----- Extracting the times of the first file ----- #
    for file in files: #loops through all the files in the main folder
        if ((file.find('Deg_#1.DTA')!=-1) or (file.find('Deg__#1.DTA')!=-1)) and (file.find('OCV')!=-1): #Finding the first file of the test, this is the OCV file at the start of the cycle
            file1 = os.path.join(folder_loc,file) #creates full file path
            T0_stamp = fl.get_timestamp(file1) #gets time stamp from first file
            t0 = int(T0_stamp.strftime("%s")) #Converting Datetime to seconds from Epoch and converts to an integer

    'Creating a DataFrame and adding an excel data sheet (or creating the file if need be)'
    excel_name = '_' + cell_name + '_Data.xlsx'
    excel_file = os.path.join(folder_loc,excel_name)
    sheet_name = 'Bias Resistance Deg'
    bias_deg_data = False

    if os.path.exists(excel_file)==True: # Looking for the data excel file in the button cell folder
        writer = pd.ExcelWriter(excel_file,engine='openpyxl',mode='a') #Creates a Pandas Excel writer using openpyxl as the engine in append mode
        wb = load_workbook(excel_file, read_only=True) # Looking for the bias Deg eis
        if sheet_name in wb.sheetnames:
            bias_deg_data = True
    elif os.path.exists(excel_file)==False:
        writer = pd.ExcelWriter(excel_file,engine='xlsxwriter') #Creates a Pandas Excel writer using XlsxWriter as the engine. This will make a new excel file

    if bias_deg_data == False:
        # ----- Extracting Ohmic, Rp, and time for each fit ----- #
        bias_fits = [f for f in os.listdir(jar_bias) if not f.startswith('.')] # Basically OS.listdir, but it excludes the hidden files (like .DS_Store)
        bias_fits = natsort.humansorted(bias_fits,key=lambda y: (len(y),y)) #This list comprehension came from stacked overflow
        
        bias_info = [] #initializing list to store the ohmic,rp, and time of the eis fits under bias

        for fit in bias_fits:
            # --- Initializing the inverter:
            inv = Inverter() #Initializes inverter object
            inv.load_fit_data(os.path.join(jar_bias,fit))

            # --- Extracting data from the fits
            ohmic = inv.R_inf*area #Ohm*cm^2
            rp = inv.predict_Rp()*area #Ohm*cm^2
            rtot = ohmic+rp #Ohm*cm^2

            # --- Extracting the time values from the original EIS files NOT the fits
            identifier = fit[fit.find('_Bias')+len('_Bias'):fit.rfind('.pkl')]
            for file in peis_deg_files:
                if (file.find('Bias')!=-1) and (file.find(identifier)!=-1): 
                    #searching through all relevant EIS files in the main folder, the first statements narrow down the bias EIS files
                    #The last statement identifies the file of choice
                    file_time = int(fl.get_timestamp(os.path.join(folder_loc,file)).strftime("%s")) #Extract time of file, converts to seconds from epoch, and converts from a str to an int
                    deg_time = (file_time-t0)/3600 # (hrs) subtracts from start time to get time from start of deg test in hours
                    break
            
            info = (ohmic,rp,rtot,deg_time) #creates a tuple of the useful information
            bias_info.append(info) #appends newly made tuple to list of all information

        df = pd.DataFrame(bias_info, columns=['ohmic', 'rp', 'rtot','time (hrs)'])
        df.to_excel(writer, sheet_name=sheet_name, index=False) # Writes this DataFrame to a specific worksheet
        writer.save() # Close the Pandas Excel writer and output the Excel file.

    elif bias_deg_data == True:
        df = pd.read_excel(excel_file,sheet_name)

    # ----- Plotting ----- #
    fig,ax = plt.subplots()
    plt.xlabel('Time (hrs)',size='x-large')
    plt.ylabel('Resistance (\u03A9 cm$^2$)',size='x-large')
    ax.plot(df['time (hrs)'],df['ohmic'],'.',color = '#21314D',label='Ohmic',ms=14) #plots ohmic resistance
    ax.plot(df['time (hrs)'],df['rp'],'.',color = '#D2492A',label='Rp',ms=14) #plots Polarization resistance
    ax.plot(df['time (hrs)'],df['rtot'],'.',color = '0.35',label='Total',ms=14) #plots total resistance
    ax.tick_params(axis='both', which='major', labelsize='large')
    if title == True: #if titles are desired one will be printed on the graph
        plt.title('Evolution under bias',size='x-large')
    plt.tight_layout() #keeps everything in the figure frame
    ax.set_ylim(ymin=0) #Ensures the Y values on the graph start at 0
    plt.legend() #Shows Legend
    plt.show()

    'Extracting and plotting ocv results (if acquired and desired)'
    if plot_ocv == True:
        # ----- Extracting Ohmic, Rp, and time for each fit ----- #
        ocv_fits = [f for f in os.listdir(jar_ocv) if not f.startswith('.')] # Basically OS.listdir, but it excludes the hidden files (like .DS_Store)
        ocv_info = [] #initializing list to store the ohmic,rp, and time of the eis fits under bias

        'Creating a DataFrame and adding an excel data sheet (or creating the file if need be)'
        excel_name = '_' + cell_name + '_Data.xlsx' #Same as before
        excel_file = os.path.join(folder_loc,excel_name) #Same as before
        sheet_name = 'OCV Resistance Deg'

        exists, writer = excel_datasheet_exists(excel_file,sheet_name)

        if exists == False:
            # ----- Extracting Ohmic, Rp, and time for each fit ----- #
            ocv_fits = [f for f in os.listdir(jar_ocv) if not f.startswith('.')] # Basically OS.listdir, but it excludes the hidden files (like .DS_Store)
            ocv_fits = natsort.humansorted(ocv_fits,key=lambda y: (len(y),y))
        
            #The above list comprehension came from stacked overflow
            bias_info = [] #initializing list to store the ohmic,rp, and time of the eis fits under bias
            for fit in ocv_fits:
                # --- Initializing the inverter:
                inv = Inverter() #Initializes inverter object
                inv.load_fit_data(os.path.join(jar_ocv,fit))

                # --- Extracting data from the fits
                ohmic = inv.R_inf*area # Ohm*cm^2
                rp = inv.predict_Rp()*area # Ohm*cm^2
                rtot = ohmic+rp # Ohm*cm^2

                # --- Extracting the time values from the original EIS files NOT the fits
                identifier = fit.split('_OCV')[1].split('.pkl')[0]
                for file in files:
                    if (file.find('PEIS')!=-1) and (file.find('_Deg')!=-1) and (file.find('Bias')==-1) and (file.find('.DTA')!=-1) and (file.find(identifier)!=-1): 
                        #searching through all relevant EIS files in the main folder, the first 4 statements narrow down the search to relevant EIS files
                        #The last statement identifies the file of choice
                        file_time = int(fl.get_timestamp(os.path.join(folder_loc,file)).strftime("%s")) #Extract time of file, converts to seconds from epoch, and converts from a str to an int
                        deg_time = (file_time-t0)/3600 # (hrs) subtracts from start time to get time from start of deg test in hours
                        break
                
                info = (ohmic,rp,rtot,deg_time) #creates a tuple of the useful information
                ocv_info.append(info) #appends newly made tuple to list of all information
            
            df = pd.DataFrame(ocv_info, columns=['ohmic', 'rp', 'rtot', 'time (hrs)'])
            df.to_excel(writer, sheet_name=sheet_name, index=False) # Writes this DataFrame to a specific worksheet
            writer.save() # Close the Pandas Excel writer and output the Excel file.

        elif bias_deg_data == True:
            df = pd.read_excel(excel_file,sheet_name)

        # ----- Plotting ----- #
        fig,ax = plt.subplots()
        plt.xlabel('Time (hrs)',size='x-large')
        plt.ylabel('Resistance (\u03A9 cm$^2$)',size='x-large')
        ax.plot(df['time (hrs)'],df['ohmic'],'.',color = '#21314D',label='Ohmic',ms=14) #plots ohmic resistance
        ax.plot(df['time (hrs)'],df['rp'],'.',color = '#D2492A',label='Rp',ms=14) #plots Polarization resistance
        ax.plot(df['time (hrs)'],df['rtot'],'.',color = '0.35',label='Total',ms=14) #plots total resistance
        ax.tick_params(axis='both', which='major', labelsize='large')
        if title == True: #if titles are desired one will be printed on the graph
            plt.title('Evolution under OCV',size='x-large')
        plt.tight_layout() #keeps everything in the figure frame
        ax.set_ylim(ymin=0) #Ensures the Y values on the graph start at 0
        plt.legend() #Shows Legend
        plt.show()

def standard_performance(loc:str, jar:str, area:float=0.5, **peis_args):
    '''
    Plots the EIS, map-fits and plots the DRT, and prints the ohmic and polarization resistance for a cell
    Generally this function is used on the first EIS spectra taken at standard testing conditions

    Parameters
    ----------
    loc, str: (path to a directory)
        The location of the folder containing the EIS files (path to folder)
    jar, str: (path to a directory)
        The location of the jar containing the map-fits (path to jar)
    area, float: 
        The active cell area in cm^2 (default = 0.5)
    peis_args, dict: 
        Any additional arguments to be passed to the plot_peis function

    Return --> None, but it plots EIS and DRT. It also prints the ohmic and rp values of the cell
    '''
    cell_name = os.path.basename(loc).split("_", 1)[0] #gets the name of the cell

    # --- Fitting DRT, and making sure I am not re-fitting it if it has already been fit
    pickle_jar = os.listdir(jar)
    pickle_name = 0
    fit_name = cell_name+'_map_fit_standard.pkl'
    for pickle in pickle_jar: # checks to see if this has already been fit, if so name gets set to 1
        if fit_name == pickle:
            pickle_name = pickle_name + 1
            break

    if pickle_name == 0:
        map_drt_save(loc,jar,fit_name)

    # ---- Loading DRT and plotting
    inv = Inverter()
    inv.load_fit_data(os.path.join(jar,fit_name))
    
    fig, ax = plt.subplots()
    bp.plot_distribution(None,inv,ax,unit_scale='',label=cell_name)
    ax.legend()
    
    plt.tight_layout()
    plt.show()

    # ---- Plotting EIS
    ohmic = inv.R_inf*area
    rp = inv.predict_Rp()*area
    ohmic_rtot = [round(float(ohmic),2),round(float(rp),2)+round(float(ohmic),2)]
    plot_peis(area,loc,**peis_args) #Plots standard PEIS spectra
    plt.show()
    print('Standard Ohmic',round(ohmic,3),'\u03A9 cm$^2^')
    print('Standard Rp',round(rp,3),'\u03A9 cm$^2^')

def po2_plots(folder_loc:str, fit_path:str, area:float, eis:bool=True, drt:bool=True,
                 o2_dependence:bool=True, drt_peaks:bool=True, print_resistance_values:bool=False,
                  ncol:int=1, legend_loc:str='best', flow100:bool = False, cut_inductance:bool = False):
    '''
    Searches through the folder_loc for all the changes in O2 concentration EIS files.
    Map-fits all of hte eis files and saves them to fit_path. If the files are already fit, they will not be re-fit.
    Plots the EIS for each concentration in one plot if eis=True and the DRT of each concentration in another if drt=True
    If O2_dependence = True, plots the ln(1/ASR) vs. ln(O2 concentration) for the ohmic and polarization resistance

    Parameters:
    -----------
    folder_loc, str: (path to folder)
        The location of the folder containing the EIS files 
    fit_path, str: (path to folder)
        The location of the folder to save the map-fits
    area, float:
        The active cell area in cm^2
    eis, bool: (default = True)
        If True, plots the EIS for each concentration in one plot
    drt, bool: (default = True)
        If True, plots the DRT for each concentration in another plot
    o2_dependence, bool: (default = True)
        If True, plots the ln(1/ASR) vs. ln(O2 concentration) for the ohmic and polarization resistance
    drt_peaks, bool: (default = True)
        If True, the DRT peaks are fit and plotted. If the peaks have not been fit, they
        will be fit and added the the cell data excel sheet. If the data exists in the spreadsheet already
        then that data will be plotted.
    print_resistance_values, bool: (default = False)
        If True, prints the ohmic and polarization resistance for each concentration
    ncol, int: (default = 1)
        The number of columns to use for the plot legend 
    legend_loc, str: 
        The location of the legend (default = 'best'). The other option is to put 
        the legend outside the figure and this is done by setting legend_loc = 'outside'
    flow100, bool: (default = False)
        Set this to true if the total flow rate for the PO2 test was 100 SCCM
        for my older tests (cell 16 and below) my total flow rate was 100 SCCM
        This just changes the list of strings to look for
    cut_inductance, bool: (default = False)
        If this is set to true, the negative inductance values at the beginning of the DataFrame
        
    Return --> None, but it plots and shows one or more plots
    '''
    
    # +++++---- Finding correct files and formatting ----+++++ #
    dta_files = [file for file in os.listdir(folder_loc) if file.endswith('.DTA')] #Makes a list of all .DTA files in the folder loc
    substring_list = ['PEIS_20O2.80Ar','PEIS_40O2.60Ar','PEIS_60O2.40Ar','PEIS_80O2.20Ar','PEIS_100O2.0Ar'] # setting variables to look for
    # substring_list = ['_20O2.80Ar','_40O2.60Ar','_60O2.40Ar','_80O2.20Ar','_100O2.0Ar'] # setting variables to look for

    if flow100 == True:
        substring_list = ['PEIS_10O2.40Ar','PEIS_20O2.30Ar','PEIS_30O2.20Ar','PEIS_40O2.10Ar','PEIS_50O2.0Ar']

    O2_conc_list = [dta_files for dta_files in dta_files if any(sub in dta_files for sub in substring_list)] # placing all the changes in O2 eis files into a list
    O2_conc_list = sorted(O2_conc_list, key=lambda x: int((x[x.find('PEIS_')+len('PEIS_'):x.rfind('O2')]))) #Sorts numerically by PO2
    # O2_conc_list = sorted(O2_conc_list, key=lambda x: int((x[x.find('H2O_')+len('H2O_'):x.rfind('O2')]))) #Sorts numerically by PO2

    cell_name = os.path.basename(fit_path).split("_", 2)[1]
    
    'Plotting all the PO2 concentration EIS spectra'
    if eis == True:
        for peis in O2_conc_list:
            # --- Finding and formatting
            loc = os.path.join(folder_loc, peis) # creates the full path to the file
            po2 = peis[peis.find('PEIS_')+len('PEIS_'):peis.rfind('O2')] #gets the pO2 from the file name
            # po2 = peis[peis.find('H2O_')+len('H2O_'):peis.rfind('O2')] #gets the pO2 from the file name

            nyquist_name =  po2 + '% O$_2$'

            if flow100 == True:
                po2_int = int(po2)
                po2 = str(po2_int * 2)
                nyquist_name = po2 + '% O$_2$'

            # --- Plotting
            plot_peiss(area,nyquist_name,loc,ncol=ncol,legend_loc=legend_loc,cut_inductance = cut_inductance)

        plt.show()

    'Plotting all the PO2 concentration EIS spectra'
    if drt == True: 
        fig, ax = plt.subplots() #initializing plots for DRT
        # --- Initializing lists for LnPO2
        O2_conc = np.array([]) # Concentration of Oxygen
        ohmic_asr = np.array([]) # ohm*cm^2
        rp_asr = np.array([]) # ohm*cm^2

        for peis in O2_conc_list: #For loop to plot the DRT of all PO2 files (probably a more elegant way, but this works)
            # --- Finding and formatting
            loc = os.path.join(folder_loc, peis) # creates the full path to the file
            po2 = peis[peis.find('PEIS_')+len('PEIS_'):peis.rfind('O2')] #gets the pO2 from the file name
            # po2 = peis[peis.find('H2O_')+len('H2O_'):peis.rfind('O2')] #gets the pO2 from the file name

            nyquist_name =  po2 + '% O$_2$'

            if flow100 == True:
                po2_int = int(po2)
                po2 = str(po2_int * 2)
                nyquist_name = po2 + '% O$_2$'

            # --- Plotting
            inv = Inverter()
            inv.load_fit_data(os.path.join(fit_path,cell_name + '_map_fit_' + po2 + 'pO2.pkl'))
            bp.plot_distribution(None,inv,ax,unit_scale='',label= po2+ '% O$_2$')

            # --- Appending resistance values to lists for lnPO2 plotting
            O2_conc = np.append(O2_conc,float(po2)/100)
            ohmic_asr = np.append(ohmic_asr,inv.R_inf*area)
            rp_asr = np.append(rp_asr,inv.predict_Rp()*area)

            # --- Printing the resistance values if that is desired
            if print_resistance_values == True:
                print(po2 + '% Oxygen Ohmic', round(inv.R_inf*area,3), '\u03A9 cm$^2^')
                print(po2 + '% Oxygen Rp', round(inv.predict_Rp()*area,3), '\u03A9 cm$^2^')
        
        ax.legend()
        plt.show()

        'Creating a DataFrame and adding an excel data sheet (or creating the file if need be)'
        excel_name = '_' + cell_name + '_Data.xlsx'
        excel_file = os.path.join(folder_loc,excel_name)
        sheet_name = 'pO2'
        exists = False

        if os.path.exists(excel_file)==True: # Looking for the data excel file in the button cell folder
            writer = pd.ExcelWriter(excel_file,engine='openpyxl',mode='a') #Creates a Pandas Excel writer using openpyxl as the engine in append mode
            wb = load_workbook(excel_file, read_only=True) # Looking for the po2
            
            if sheet_name in wb.sheetnames:
                exists = True
        
        elif os.path.exists(excel_file)==False:
            writer = pd.ExcelWriter(excel_file,engine='xlsxwriter') #Creates a Pandas Excel writer using XlsxWriter as the engine. This will make a new excel file

        if exists == False:
            df_po2 = pd.DataFrame(list(zip(O2_conc*100,ohmic_asr,rp_asr)),
                columns =['O2 Concentration (%)','Ohmic ASR (ohm*cm$^2$)', 'Rp ASR (ohm*cm$^2$)'])
            df_po2.to_excel(writer, sheet_name=sheet_name, index=False) # Writes this DataFrame to a specific worksheet
            writer.save() # Close the Pandas Excel writer and output the Excel file.

    ' ------ Plotting the Oxygen dependence'
    if o2_dependence == True:
        lnpo2(ohmic_asr,rp_asr,O2_conc)
    elif drt==False and o2_dependence==True or print_resistance_values==True:
        print('Set drt to True. The cell resistance values are found using Jakes DRT package and thus the EIS spectra need to be fit')

    ' --- DRT peak fitting and plotting --- '
    if drt_peaks == True:
        o2_map_fits = [file for file in os.listdir(fit_path) if file.find('pO2')!=-1] # gets all fuel cell map fits
        o2_map_fits = natsort.humansorted(o2_map_fits,key=lambda y: (len(y),y)) #sorts by bias
        
        # --- Checking to see if the peaks have already been fit:
        peak_data = False
        excel_name = '_' + cell_name + '_Data.xlsx'
        excel_file = os.path.join(folder_loc,excel_name)
        peak_data_sheet = 'pO2 DRT peak fits'

        exists, writer = excel_datasheet_exists(excel_file,peak_data_sheet)
        
        if exists == False: # Make the excel data list
            # --- Fitting peaks and appending to a DataFrame
            df_tau_r = pd.DataFrame(columns = ['O2 Concentration (%)','Tau','Resistance']) #Initializing DataFrame to save temperature
            for fit in o2_map_fits: #Loading DRT, fitting peaks, and saving to a DataFrame
                # creating inverter and calling fits
                inv = Inverter()
                inv.load_fit_data(os.path.join(fit_path,fit))
                inv.fit_peaks(prom_rthresh=0.05) # fit the peaks

                # --- obtain time constants from inverters
                tau = inv.extract_peak_info().get('tau_0') # τ/s
                r = inv.extract_peak_info().get('R')*area # Ω*cm2

                # --- Obtaining bias
                o2_conc = int(fit[fit.find('fit_')+len('fit_'):fit.rfind('pO2')])

                # Appending tau and r for each peak into df_tau_r
                i = 0
                for τ in tau:
                    df_tau_r.loc[len(df_tau_r.index)] = [o2_conc, τ, r[i]]
                    i = i+1

            df_tau_r.to_excel(writer, sheet_name=peak_data_sheet, index=False) # Extract data to an excel sheet
            writer.save() # Close the Pandas Excel writer and output the Excel file.

        elif exists == True: #load the data into a DataFrame
            df_tau_r = pd.read_excel(excel_file,peak_data_sheet)
        
        # ----- plotting
        palette = sns.color_palette('crest_r', as_cmap=True)
        plot = sns.scatterplot(x = 'Tau', y = 'Resistance', data = df_tau_r, hue='O2 Concentration (%)',palette = palette,s=69)

        # ----- Ascetics stuff
        sns.set_context("talk")
        fontsize = 14
        sns.despine()
        plot.set_ylabel('ASR (\u03A9 cm$^2$)',fontsize=fontsize)
        plot.set_xlabel('Time Constant (\u03C4/s)',fontsize=fontsize)
        plot.set(xscale='log')
        plt.tight_layout()
        plt.show()

def fc_bias_plots(folder_loc:str, fit_path:str, area:float, eis:bool=True, drt:bool=True,
                    drt_peaks:bool=True, ncol:int=1, legend_loc:str='best'):
    '''
    Finds all EIS files in the folder_loc taken during bias testing in Fuel Cell mode and plots the EIS
    The corresponding DRT fits are located and also plotted if drt = True
    If drt_peaks = True, the DRT spectra are fit, and the resistance of each DRT peak at each condition is plotted
    All data is append to a sheet in the cell data excel file if it does not already exist. If the data
    does already exist that data is called upon for plotting

    The .DTA EIS files are taken during a Gamry sequence that I use for testing the cell performance under
    various biases

    Parameters:
    -----------
    folder_loc, str:
        The folder location of the EIS files (path to directory)
    fit_path, str:
        The folder location of the DRT fits (path to directory)
    area, float: 
        The active cell area in cm^2
    eis, bool: (default = True)
        If True, the EIS spectra are plotted
    drt, bool: (default = True)
        If True, the DRT spectra are plotted
    drt_peaks, bool: (default = True)
        If True, the DRT peaks are fit and plotted. If the peaks have not been fit, they
        will be fit and added the the cell data excel sheet. If the data exists in the spreadsheet already
        then that data will be plotted.
    ncol, int: (default = 1)
        The number of columns in the legend
    legend_loc, str: (default = 'best')
        The location of the legend. The other option is to put 
        the legend outside the figure and this is done by setting legend_loc = 'outside'

    Return --> None but one or more plots are created and shown
    '''
    
    'Finding correct files and formatting'
    dta_files = [file for file in os.listdir(folder_loc) if file.endswith('.DTA')] #Makes a list of all .DTA files in the folder loc
    bias_eis = [] #initializing bias files list
    for file in dta_files: #Finding all fuel cell bias EIS files
        if (file.find('PEIS')!=-1) and (file.find('bias.DTA')!=-1) and (file.find('_n')!=-1):
            bias_eis.append(os.path.join(folder_loc,file))
    cell_name = os.path.basename(fit_path).split("_", 2)[1]
    bias_eis = sorted(bias_eis, key=lambda x: int((x[x.find('550C_n')+len('550C_n'):x.rfind('bias')]))) #Sorts numerically by bias

    for file in dta_files: #Finding the 0 bias condition and adding it to the beginning of the list of bias_eis
        if (file.find('PEIS')!=-1) and (file.find('0bias.DTA')!=-1): #adding in the 0 bias condition
            bias_eis.insert(0,os.path.join(folder_loc,file))

    'Plotting EIS'
    if eis == True:
        # --- Setting up the color map
        cmap = plt.cm.get_cmap('plasma') #cmr.redshift 
        color_space = np.linspace(0.2,0.8,len(bias_eis)) # array from 0-1 for the colormap for plotting
        c = 0 # index of the color array

        for peis in bias_eis:
            # --- Finding and formatting
            loc = os.path.join(folder_loc, peis) # creates the full path to the file
            bias = peis[peis.find('550C_')+len('PEIS_'):peis.rfind('bias')] #gets the bias from the file name
            if len(bias) > 1:
                bias = -int(bias[1:])/10

            nyquist_name =  str(bias) + 'V'

            # --- Plotting
            color = cmap(color_space[c])
            plot_peiss(area,nyquist_name,loc,ncol=ncol,legend_loc=legend_loc,color=color)
            c = c+1
        plt.show()

    'Plotting DRT'
    if drt == True:
        fig, ax = plt.subplots() #initializing plots for DRT

        # --- Setting up the color map
        cmap = plt.cm.get_cmap('plasma') #cmr.redshift 
        color_space = np.linspace(0.2,0.8,len(bias_eis)) # array from 0-1 for the colormap for plotting
        c = 0 # index of the color array

        # --- Initializing lists of data to save
        bias_array = np.array([]) # Applied bias of the eis spectra relative to OCV
        ohmic_asr = np.array([]) # ohm*cm^2
        rp_asr = np.array([]) # ohm*cm^2

        for peis in bias_eis: # For loop to plot the DRT of all PO2 files (probably a more elegant way, but this works)
            # --- Finding and formatting
            loc = os.path.join(folder_loc, peis) # creates the full path to the file
            bias = peis[peis.find('550C_')+len('PEIS_'):peis.rfind('bias')] #gets the bias from the file name
            map_fit_name = cell_name + '_map_fit_' + bias + 'bias.pkl'
            if len(bias) > 1:
                bias = -int(bias[1:])/10
                
            label =  str(bias) + 'V'

            # --- Updating arrays and plotting
            inv = Inverter()
            inv.load_fit_data(os.path.join(fit_path,map_fit_name))
            bias_array = np.append(bias_array,float(bias))
            ohmic_asr = np.append(ohmic_asr,inv.R_inf*area)
            rp_asr = np.append(rp_asr,inv.predict_Rp()*area)
            color = cmap(color_space[c])
            bp.plot_distribution(None,inv,ax,unit_scale='',label = label, color = color)
            c = c + 1
        ax.legend()
        plt.show()
        
        'Creating a DataFrame and adding an excel data sheet (or creating the file if need be)'
        excel_name = '_' + cell_name + '_Data.xlsx'
        excel_file = os.path.join(folder_loc,excel_name)
        sheet_name = 'FC mode data'
        exists = False

        if os.path.exists(excel_file)==True: # Looking for the data excel file in the button cell folder
            writer = pd.ExcelWriter(excel_file,engine='openpyxl',mode='a') #Creates a Pandas Excel writer using openpyxl as the engine in append mode
            wb = load_workbook(excel_file, read_only=True) # Looking for the po2
            if sheet_name in wb.sheetnames:
                exists = True
        elif os.path.exists(excel_file)==False:
            writer = pd.ExcelWriter(excel_file,engine='xlsxwriter') #Creates a Pandas Excel writer using XlsxWriter as the engine. This will make a new excel file

        if exists == False:
            df_fc_bias = pd.DataFrame(list(zip(bias_array,ohmic_asr,rp_asr)),
                columns =['Applied Bias (V)','Ohmic ASR (ohm*cm^2)', 'Rp ASR (ohm*cm^2)'])
            df_fc_bias.to_excel(writer, sheet_name=sheet_name, index=False) # Writes this DataFrame to a specific worksheet
            writer.save() # Close the Pandas Excel writer and output the Excel file.

    ' --- DRT peak fitting and plotting --- '
    if drt_peaks == True:
        bias_map_fits = [file for file in os.listdir(fit_path) if (file.find('bias')!=-1) and (file.find('fit_n')!=-1 or file.find('0bias')!=-1)] # gets all fuel cell map fits
        bias_map_fits = natsort.humansorted(bias_map_fits,key=lambda y: (len(y),y)) #sorts by bias
        bias_map_fits.reverse()

        # --- Checking to see if the peaks have already been fit:
        peak_data = False
        excel_name = '_' + cell_name + '_Data.xlsx'
        excel_file = os.path.join(folder_loc,excel_name)
        peak_data_sheet = 'FC Bias DRT peak fits'

        if os.path.exists(excel_file)==True: # Looking for the data excel file in the button cell folder
            writer = pd.ExcelWriter(excel_file,engine='openpyxl',mode='a') #Creates a Pandas Excel writer using openpyxl as the engine in append mode
            wb = load_workbook(excel_file, read_only=True) # Looking for the Arrhenius Data Sheet
            if peak_data_sheet in wb.sheetnames:
                peak_data = True
        elif os.path.exists(excel_file)==False:
            writer = pd.ExcelWriter(excel_file,engine='xlsxwriter') #Creates a Pandas Excel writer using XlsxWriter as the engine. This will make a new excel file
        
        if peak_data == False: # Make the excel data list
            # --- Fitting peaks and appending to a DataFrame
            df_tau_r = pd.DataFrame(columns = ['Bias','Tau','Resistance']) #Initializing DataFrame to save temperature
            for fit in bias_map_fits: #Loading DRT, fitting peaks, and saving to a DataFrame
                # creating inverter and calling fits
                inv = Inverter()
                inv.load_fit_data(os.path.join(fit_path,fit))
                inv.fit_peaks(prom_rthresh=0.05) # fit the peaks

                # --- obtain time constants from inverters
                tau = inv.extract_peak_info().get('tau_0') # τ/s
                r = inv.extract_peak_info().get('R')*area # Ω*cm2

                # --- Obtaining bias
                number = fit[fit.find('fit_n')+len('fit_n'):fit.rfind('bias')]
                try:
                    bias = -int(number)/10
                except ValueError:
                    bias = 0

                # Appending tau and r for each peak into df_tau_r
                i = 0
                for τ in tau:
                    df_tau_r.loc[len(df_tau_r.index)] = [bias, τ, r[i]]
                    i = i+1

            df_tau_r.to_excel(writer, sheet_name=peak_data_sheet, index=False) # Extract data to an excel sheet
            writer.save() # Close the Pandas Excel writer and output the Excel file.

        elif peak_data == True: #load the data into a DataFrame
            df_tau_r = pd.read_excel(excel_file,peak_data_sheet)

        # ----- plotting
        cmap = 'plasma_r'
        palette = sns.color_palette(cmap)
        plot = sns.scatterplot(x = 'Tau', y = 'Resistance', data = df_tau_r, hue='Bias',palette = palette,s=69)

        # ----- Aesthetic stuff
        sns.set_context("talk")
        fontsize = 14
        sns.despine()
        plot.set_ylabel('ASR (\u03A9 cm$^2$)',fontsize=fontsize)
        plot.set_xlabel('Time Constant (\u03C4/s)',fontsize=fontsize)
        plot.set(xscale='log')
        plt.tight_layout()
        plt.show()

def ec_bias_plots(folder_loc:str, fit_path:str, area:float, eis:bool=True,
                drt:bool=True, ncol:int=1, legend_loc:str='best'):
    '''
    Finds all EIS files in the folder_loc taken during bias testing in Electrolysis cell mode and plots the EIS
    The corresponding DRT fits are located and also plotted if drt = True
    All data is append to a sheet in the cell data excel file if it does not already exist. If the data
    does already exist that data is called upon for plotting

    The .DTA EIS files are taken during a Gamry sequence that I use for testing the cell performance under
    various biases

    param folder_loc, str: The folder location of the EIS files (path to directory)
    param fit_path, str: The folder location of the DRT fits (path to directory)
    param area, float: The active cell area in cm^2
    param eis, bool: If True, the EIS spectra are plotted
    param drt, bool: If True, the DRT spectra are plotted
    param ncol, int: The number of columns in the legend
    param legend_loc, str: The location of the legend (default = 'best'). The other option is to put 
        the legend outside the figure and this is done by setting legend_loc = 'outside'

    Return --> None but one or more plots are created and shown
    '''
    
    'Finding correct files and formatting'
    dta_files = [file for file in os.listdir(folder_loc) if file.endswith('.DTA')] #Makes a list of all .DTA files in the folder loc
    bias_eis = [] #initializing bias files list

    for file in dta_files: #Finding all fuel cell bias EIS files
        if (file.find('PEIS')!=-1) and (file.find('bias.DTA')!=-1) and (file.find('_n')==-1):
            bias_eis.append(os.path.join(folder_loc,file))

    cell_name = os.path.basename(fit_path).split("_", 2)[1]
    bias_eis = sorted(bias_eis, key=lambda x: int((x[x.find('550C_')+len('550C_'):x.rfind('bias')]))) #Sorts numerically by bias

    'Plotting EIS'
    if eis == True:
        for peis in bias_eis:
            # --- Finding and formatting
            loc = os.path.join(folder_loc, peis) # creates the full path to the file
            bias = peis[peis.find('550C_')+len('PEIS_'):peis.rfind('bias')] #gets the bias from the file name
            bias = int(bias)/10

            nyquist_name =  str(bias) + 'V'

            # --- Plotting
            plot_peiss(area,nyquist_name,loc,ncol=ncol,legend_loc=legend_loc)
        plt.show()

    'Plotting DRT'
    if drt == True:
        fig, ax = plt.subplots() #initializing plots for DRT
        # --- Initializing lists of data to save
        bias_array = np.array([]) # Applied bias of the eis spectra relative to OCV
        ohmic_asr = np.array([]) # ohm*cm^2
        rp_asr = np.array([]) # ohm*cm^2
        for peis in bias_eis: # For loop to plot the DRT of all PO2 files (probably a more elegant way, but this works)
            # --- Finding and formatting
            loc = os.path.join(folder_loc, peis) # creates the full path to the file
            bias = peis[peis.find('550C_')+len('PEIS_'):peis.rfind('bias')] #gets the bias from the file name
            map_fit_name = cell_name + '_map_fit_' + bias + 'bias.pkl'
            if int(bias) > 0:
                bias = int(bias)/10

            label =  str(bias) + 'V'

            # --- Updating arrays and Plotting
            inv = Inverter()
            inv.load_fit_data(os.path.join(fit_path,map_fit_name))
            bias_array = np.append(bias_array,float(bias))
            ohmic_asr = np.append(ohmic_asr,inv.R_inf*area)
            rp_asr = np.append(rp_asr,inv.predict_Rp()*area)
            bp.plot_distribution(None,inv,ax,unit_scale='',label = label)
        ax.legend()
        plt.show()

        'Creating a DataFrame and adding an excel data sheet (or creating the file if need be)'
        excel_name = '_' + cell_name + '_Data.xlsx'
        excel_file = os.path.join(folder_loc,excel_name)
        sheet_name = 'EC mode data'
        exists = False

        if os.path.exists(excel_file)==True: # Looking for the data excel file in the button cell folder
            writer = pd.ExcelWriter(excel_file,engine='openpyxl',mode='a') #Creates a Pandas Excel writer using openpyxl as the engine in append mode
            wb = load_workbook(excel_file, read_only=True) # Looking for the po2
            if sheet_name in wb.sheetnames:
                exists = True
        elif os.path.exists(excel_file)==False:
            writer = pd.ExcelWriter(excel_file,engine='xlsxwriter') #Creates a Pandas Excel writer using XlsxWriter as the engine. This will make a new excel file

        if exists == False:
            df_ec_bias = pd.DataFrame(list(zip(bias_array,ohmic_asr,rp_asr)),
                columns =['Applied Bias (V)','Ohmic ASR (ohm*cm^2)', 'Rp ASR (ohm*cm^2)'])
            df_ec_bias.to_excel(writer, sheet_name=sheet_name, index=False) # Writes this DataFrame to a specific worksheet
            writer.save() # Close the Pandas Excel writer and output the Excel file.

def steam_plots(folder_loc:str, fit_path:str, area:float, legend_loc:bool='best', O2:int=40,
                which:str = 'core', init_from_ridge:bool=True, **peis_args:dict):
    '''
    Searches through folder_loc for the steam plot EIS spectra (they will have certain names to aid this function),
    and plots the EIS. Also the DRT is fit, saved, and plotted.

    The ohmic and Rp data at each condition is saved to the cell data excel file if the sheet 
    does not already exist

    Parameters:
    -----------
    folder_loc, str: 
        The folder location of the EIS files (path to directory)
    fit_path, str: 
        The folder location of the DRT fits (path to directory)
    area, float: 
        The active cell area in cm^2
    legend_loc, str: 
        The location of the legend (default = 'best'). The other option is to put 
        the legend outside the figure and this is done by setting legend_loc = 'outside'
    O2, int: 
        The O2 flow rate in SCCM used when taking the EIS spectra and in naming the files (default = 40)
    which, string: 
        which data to store. 'core' or 'sample'. Core file sizes are smaller
    init_from_ridge, bool: optional (default: False)
        If True, use the hyper-parametric ridge solution to initialize the Bayesian fit.
        Only valid for single-distribution fits
    peis_args, dict: (optional)
        passes arguments to the plot_peis function

    Return --> None but one or more plots are generated
    '''
    
    'Finding and formatting'
    # --- compiling a list of all the eis files associated with steam concentration testing
    cell_name = os.path.basename(fit_path).split("_", 2)[1]
    eis_files = [file for file in os.listdir(folder_loc) if file.endswith('.DTA') and file.find('PEIS')!=-1] #Makes a list of all .DTA files in the folder loc
    steam_eis = [] #initializing bias files list
    for file in eis_files: #Finding all steam EIS files
        if file.find(str(O2)+'O2')!=-1 and (file.find('60Ar')==-1):
            steam_eis.append(os.path.join(folder_loc,file))
        elif (file.find('Steam')!=-1) or (file.find('steam')!=-1):
            steam_eis.append(os.path.join(folder_loc,file))

    # --- Sorting said list
    steam_eis = natsort.humansorted(steam_eis,reverse=True)
    for eis in steam_eis:
        if (eis.find('After')!=-1) or (eis.find('after')!=-1):
            after = eis
            steam_eis.append(steam_eis.pop(steam_eis.index(after)))

    '# --- Map fitting EIS if not already done so'
    pickle_jar = os.listdir(fit_path)
    pickle_name = 0
    map_fits =[] #stores the DRT map fits to save and call
    for eis in steam_eis:
        if eis.find('40O2')!=-1:
            steam_percent = eis[eis.find('Ar.')+len('Ar.'):eis.rfind('H2O')]
            fit_name = cell_name + '_map_fit_' + steam_percent + 'pH2O.pkl'
            map_fits.append(fit_name)
            for pickle in pickle_jar: # checks to see if this has already been fit, if so name gets set to 1
                if fit_name == pickle:
                    pickle_name = pickle_name + 1
                    break
            if pickle_name == 0:
                map_drt_save(eis,fit_path,fit_name,which=which,init_from_ridge=init_from_ridge)

        if (eis.find('Steam')!=-1) or (eis.find('steam')!=-1):
            indicator = eis[eis.find('550C_')+len('550C_'):eis.rfind('.DTA')]
            fit_name = cell_name + '_map_fit_' + indicator + '.pkl'
            map_fits.append(fit_name)
            for pickle in pickle_jar: # checks to see if this has already been fit, if so name gets set to 1
                if fit_name == pickle:
                    pickle_name = pickle_name + 1
                    break
            if pickle_name == 0: 
                map_drt_save(eis,fit_path,fit_name,which=which,init_from_ridge=init_from_ridge)

    '# --- Plotting EIS'
    for eis in steam_eis:
        if eis.find('40O2')!=-1:
            condition = eis[eis.find('Ar.')+len('Ar.'):eis.rfind('H2O')] + '% H$_2$O'
        elif (eis.find('Steam')!=-1) or (eis.find('steam')!=-1):
                if (eis.find('After')!=-1) or (eis.find('after')!=-1):
                    condition = '3% H$_2$O after testing'
                elif (eis.find('b4')!=-1) or (eis.find('B4')!=-1) or (eis.find('before')!=-1) or (eis.find('Before')!=-1):
                    condition = '3% H$_2$O before testing'
        plot_peiss(area,condition,eis,legend_loc=legend_loc,**peis_args)
        # plt.ylim(-0.1,0.05)
    plt.show() 

    '# --- Plotting DRT'
    fig, ax = plt.subplots() #initializing plots for DRT
    # --- Initializing lists of data to save
    steam_array = np.array([]) # Concentration of steam/before or after
    ohmic_asr = np.array([]) # ohm*cm^2
    rp_asr = np.array([]) # ohm*cm^2
    for fit in map_fits:
        label = fit[fit.find('_fit_')+len('_fit_'):fit.rfind('.pkl')]
        inv = Inverter()
        inv.load_fit_data(os.path.join(fit_path,fit))
        steam_array = np.append(steam_array,label)
        ohmic_asr = np.append(ohmic_asr,inv.R_inf*area)
        rp_asr = np.append(rp_asr,inv.predict_Rp()*area)
        bp.plot_distribution(None,inv,ax,unit_scale='',label = label)
    ax.legend()
    plt.show()

    'Creating a DataFrame and adding an excel data sheet (or creating the file if need be)'
    excel_name = '_' + cell_name + '_Data.xlsx'
    excel_file = os.path.join(folder_loc,excel_name)
    sheet_name = 'Steam data'
    exists = False

    if os.path.exists(excel_file)==True: # Looking for the data excel file in the button cell folder
        writer = pd.ExcelWriter(excel_file,engine='openpyxl',mode='a') #Creates a Pandas Excel writer using openpyxl as the engine in append mode
        wb = load_workbook(excel_file, read_only=True) # Looking for the po2
        if sheet_name in wb.sheetnames:
            exists = True
    elif os.path.exists(excel_file)==False:
        writer = pd.ExcelWriter(excel_file,engine='xlsxwriter') #Creates a Pandas Excel writer using XlsxWriter as the engine. This will make a new excel file

    if exists == False:
        df_ec_bias = pd.DataFrame(list(zip(steam_array,ohmic_asr,rp_asr)),
            columns =['Steam Condition','Ohmic ASR (ohm*cm^2)', 'Rp ASR (ohm*cm^2)'])
        df_ec_bias.to_excel(writer, sheet_name=sheet_name, index=False) # Writes this DataFrame to a specific worksheet
        writer.save() # Close the Pandas Excel writer and output the Excel file.

def deg_ocv_eis_plots(folder_loc:str, jar_loc:str, area:float, start_file:str = 'default',
                        a10:bool=True, eis:bool=True, drt:bool=True, ncol:int=1,
                        legend_loc:str='outside', cbar = True, rev = True):
    '''
    Plots the EIS and the fit DRT of a cell taken at OCV during a long term stability test. This function
    complements a Gamry sequence that I use to test the cell stability over time in fuel cell mode. 
    DRT must already be fit before using this function.

    OCV eis spectra are taken one an hour for the first 10 hours then 1 every 10 hours for the rest of the test
    
    Parameters
    ----------
    folder_loc, str: (path to a directory)
        The folder location of the EIS files 
    jar_loc, str: (path to a directory)
        The folder location of the pickle jar where the DRT fits are 
    area, float: 
        The active cell area in cm^2
    start_file, str: (default = 'default')
        The name of the first file taken in the stability sequence
    a10, bool: (After 10 hours) (default = True)
        Whether or not to plot the first 10 hours of the test.
        if True, then only hours 10 onward are plotted
    eis, bool: (default = True)
        Whether or not to plot the EIS spectra
    drt, bool: (default = True)
        Whether or not to plot the DRT fits
    ncol, int: (default = 1)
        The number of columns to use in the legend of the plot
    legend_loc, str:  (default = 'outside')
        The location of the legend (default = 'best'). The other option is to put 
        the legend outside the figure and this is done by setting legend_loc = 'outside'
    cbar, Bool: (default = True)
        If this is set to true then the legend will just be a colorbar
        If false, then the normal legend will exist
    rev, Bol: (Default = True)
        Reverses the order that the DRT is shown in. This needs to be a variable because
        a10_ocv_fits and a10_ocv_deg_peis_time need to move in tandem or the legend will be off

    Return --> None but one or more plots are created and shown
    '''
    
    # cell_name = os.path.basename(folder_loc).split("_", 3)[2]


    if start_file == 'default': #if another file is specified as the first file, this file will be used to find T0
        for file in os.listdir(folder_loc): #Finding the first file
            if file.find('Deg__#1.DTA')!=-1 and file.find('OCV')!=-1:
                start_file = os.path.join(folder_loc,file)
        t0 = int(fl.get_timestamp(start_file).strftime("%s")) # Getting time stamp for first file in s from epoch, and convert to int
    else:
        t0 = int(fl.get_timestamp(start_file).strftime("%s")) # Getting time stamp for first file in s from epoch, and convert to int

    # cell_name = os.path.basename(folder_loc).split("_", 3)[2]

    for file in os.listdir(folder_loc): # Finding the name of the cell
        if file.find('PEIS')!=-1:
            cell_name = os.path.basename(file).split("_", 1)[0]
            break

    if eis == True:
        'Finding correct EIS files and formatting'
        dta_files = [file for file in os.listdir(folder_loc) if file.endswith('.DTA')] #Makes a list of all .DTA files in the folder loc
        ocv_deg_peis = [] #initializing bias files list
        
        if a10 == False: #Plot the EIS for the first 10 hours in 1 hr increments (the increment it is taken in)
            for file in dta_files: #Finding all fuel cell bias EIS files
                if (file.find('PEIS')!=-1) and (file.find('_Deg')!=-1) and (file.find('Deg10')==-1) and (file.find('n3Bias')==-1):
                    ocv_deg_peis.append(os.path.join(folder_loc,file))

            f10_ocv_deg_peis = sorted(ocv_deg_peis, key=lambda x: int((x[x.find('__#')+len('__#'):x.rfind('.DTA')]))) #Sorts numerically by time
            
            'Plotting EIS'
            for peis in f10_ocv_deg_peis:
                # --- Finding and formatting
                loc = os.path.join(folder_loc, peis) # creates the full path to the file
                time = peis[peis.find('__#')+len('__#'):peis.rfind('.DTA')] #gets the bias from the file name
                
                nyquist_name =  str(time) + ' Hours'

                # --- Plotting
                plot_peiss(area,nyquist_name,loc,ncol=ncol,legend_loc=legend_loc)

            plt.show()

        if a10 == True: # Plot the DRT after 10 hours in 10 hour increments (the increment it is taken in)
            'Finding and formatting'
            for file in dta_files: #Finding all fuel cell bias EIS files
                if (file.find('PEIS')!=-1) and (file.find('_Deg10')!=-1) and (file.find('n3Bias')==-1):
                    loc = os.path.join(folder_loc,file)
                    ocv_deg_peis.append((loc,int(fl.get_timestamp(loc).strftime("%s"))))
            
            'Sorting values by time'
            ocv_deg_peis.sort(key=lambda x:x[1])
            a10_ocv_deg_peis = ocv_deg_peis

            # --- Getting the ent time
            last = a10_ocv_deg_peis[-1][1]
            end_time = int((last-t0)/3600) #hrs

            'Setting up an array for the color map'
            # color = np.linspace(0,1,len(a10_ocv_deg_peis)) # array from 0-1 for the colormap for plotting
            cmap = plt.cm.get_cmap('viridis', end_time)

            'Plotting EIS'
            fig,ax = plt.subplots()
            for peis in a10_ocv_deg_peis:
                loc = os.path.join(folder_loc, peis[0]) # creates the full path to the file

                # --- Finding time of the EIS from the start of the degradation test
                test_time = peis[1]
                time = round((test_time-t0)/3600) #hrs
                nyquist_name =  str(time) + ' Hours'

                # --- Plotting
                # plot_peiss(area,nyquist_name,loc,ncol=ncol,legend_loc=legend_loc,color = cmap(color[c]))
                df_useful = peis_data(area,loc)
                if cbar == True:
                    ax.plot(df_useful['ohm'],df_useful['ohm.1'],'o',markersize=9,label = nyquist_name,color = cmap(time)) #plots data

                if cbar == False:
                    plot_peiss(area,nyquist_name,loc,ncol=ncol,legend_loc=legend_loc,color = cmap(time))


                # --- Plot Formatting                  
                ax.set_xlabel('Zreal (\u03A9 cm$^2$)',fontsize='xx-large')
                ax.set_ylabel('$\u2212$Zimag (\u03A9 cm$^2$)',fontsize='xx-large')
                ax.tick_params(axis='both', which='major', labelsize='x-large')
                ax.axhline(y=0,color='k', linestyle='-.') # plots line at 0 #D2492A is the color of Mines orange
                ax.axis('scaled') #Keeps X and Y axis scaled 1 to 1
                ax.spines['top'].set_visible(False)
                ax.spines['right'].set_visible(False)

            if cbar == True:
                sm = plt.cm.ScalarMappable(cmap=cmap)
                sm.set_array([0,end_time])
                divider = make_axes_locatable(ax)
                cax = divider.append_axes("right", size="5%", pad=0.05)
                cb = plt.colorbar(sm,ticks = [0,end_time],cax=cax)
                cb.set_label(label='Time (hrs)',fontsize = 'xx-large',labelpad = -10)
                cb.set_ticks(ticks = [0,end_time],fontsize='xx-large')
                cb.ax.tick_params(labelsize='x-large')

            plt.tight_layout()
            plt.show()

    if drt == True:
        'Finding and Formatting'
        fig, ax = plt.subplots() #initializing plots for DRT

        if a10 == False: #Plot the DRT for the first 10 hours in 1 hr increments (the increment it is taken in)
            f10_ocv_deg_fits = [file for file in os.listdir(jar_loc) if (file.find('OCV')!=-1 and file.find('OCV10')==-1)] #Makes a list of all OCV files 
            f10_ocv_deg_fits = sorted(f10_ocv_deg_fits, key=lambda x: int((x[x.find('__#')+len('__#'):x.rfind('.pkl')]))) #Sorts numerically by time
            
            for fit in f10_ocv_deg_fits: # Calling fits and plotting
                # --- Finding and formatting
                loc = os.path.join(folder_loc, fit) # creates the full path to the file
                time = fit[fit.find('__#')+len('__#'):fit.rfind('.pkl')] #gets the bias from the file name
                map_fit_name = cell_name+'_OCV__#' + time + '.pkl'
                label = str(time) + ' Hours'

                # --- Plotting
                inv = Inverter()
                inv.load_fit_data(os.path.join(jar_loc,map_fit_name))
                bp.plot_distribution(None,inv,ax,unit_scale='',label = label)

            ax.legend()
            plt.show()
        
        if a10 == True:  
            'Sorting out relevant DRT files'         
            a10_ocv_fits = [file for file in os.listdir(jar_loc) if file.find('OCV10')!=-1] #Makes a list of all OCV files 
            a10_ocv_fits = natsort.humansorted(a10_ocv_fits,reverse = rev)

            "Finding the time from the correct EIS file"
            dta_files = [file for file in os.listdir(folder_loc) if file.endswith('.DTA')] #Makes a list of all .DTA files in the folder loc
            ocv_deg_peis_time = [] #initializing bias files list
            for file in dta_files: #Finding all fuel cell ocv EIS files
                if (file.find('PEIS')!=-1) and (file.find('_Deg10')!=-1) and (file.find('n3Bias')==-1):
                    loc = os.path.join(folder_loc,file)
                    ocv_deg_peis_time.append(fl.get_timestamp(loc).strftime("%s"))

            'Sorting values by time'
            # The order of the list is reversed to ensure that if there is degradation, the DRT plot can fit the larger spectra
            a10_ocv_deg_peis_time = natsort.humansorted(ocv_deg_peis_time)
            a10_ocv_deg_peis_time.sort(reverse=rev)

            'Setting up an array for the color map'
            if rev == True:
                cmap = plt.cm.get_cmap('viridis')
            if rev == False:
                cmap = plt.cm.get_cmap('viridis_r')
                
            color_space = np.linspace(0,1,len(a10_ocv_fits)) # array from 0-1 for the colormap for plotting
            c = 0 # index of the color array

            i = 0 # Setting the initial index
            for fit in a10_ocv_fits: #For loop to plot the DRT of all PO2 files (probably a more elegant way, but this works)
                # --- Finding the right mapfit file and matching it to the right time
                number = fit[fit.find('OCV10')+len('OCV10'):fit.rfind('.pkl')] #gets the OCV (which is actually a string)
                map_fit_name = cell_name + '_OCV10' + number + '.pkl'
                time = round((int(a10_ocv_deg_peis_time[i])-t0)/3600) # to convert to hours and round to the nearest hour from the test start
                label = str(time) + ' Hours'

                i = i+1

                # --- Plotting
                inv = Inverter()
                inv.load_fit_data(os.path.join(jar_loc,map_fit_name))
                color = cmap(color_space[len(color_space)-c-1])
                bp.plot_distribution(None,inv,ax,unit_scale='',label = label,color=color)
                c = c + 1

            ax.legend()
            plt.show()

def deg_bias_eis_plots(folder_loc:str, jar_loc:str, area:float, start_file:str = 'default',
                        a10:bool=True, eis:bool=True, drt:bool=True, ncol:bool=1, 
                        legend_loc:str='outside', cbar = True):
    '''
    Plots the EIS and the fit DRT of a cell taken at bias during a long term stability test. This function
    complements a Gamry sequence that I use to test the cell stability over time in fuel cell mode. 
    DRT must already be fit before using this function.

    Bias eis spectra are taken one an hour for the first 10 hours then 1 every 10 hours for the rest of the test

    Parameters
    ----------
    folder_loc, str: (path to a directory)
        The folder location of the EIS files
    jar_loc, str:  (path to a directory)
        The folder location of the pickle jar where the DRT fits are stored
    area, float: 
        The active cell area in cm^2
    start_file, str: (default = 'default')
        The name of the first file taken in the stability sequence. If 'default' is selected, then
        the function will find the first ocv file taken at the start of the stability sequence
    a10, bool: (After 10 hours) (default = True)
        Whether or not to plot the first 10 hours of the test.
        if True, then only hours 10 onward are plotted
    eis, bool: (default = True)
        Whether or not to plot the EIS spectra
    drt, bool: (default = True)
        Whether or not to plot the DRT fits
    ncol, int: (default` = 1)
        The number of columns to use in the legend of the plot
    legend_loc, str: (default = 'outside')
        The location of the legend (default = 'best'). The other option is to put 
        the legend outside the figure and this is done by setting legend_loc = 'outside'
    cbar, Bool: (default = True)
        If this is set to true then the legend will just be a colorbar
        If false, then the normal legend will exist

    Return --> None but one or more plots are created and shown
    '''
    
    if start_file == 'default': #if another file is specified as the first file, this file will be used to find T0
        for file in os.listdir(folder_loc): #Finding the first file
            if file.find('Deg__#1.DTA')!=-1 and file.find('OCV')!=-1:
                start_file = os.path.join(folder_loc,file)
        t0 = int(fl.get_timestamp(start_file).strftime("%s")) # Getting time stamp for first file in s from epoch, and convert to int
    else:
        t0 = int(fl.get_timestamp(start_file).strftime("%s")) # Getting time stamp for first file in s from epoch, and convert to int


    # cell_name = os.path.basename(folder_loc).split("_", 3)[2]

    for file in os.listdir(folder_loc): # Finding the name of the celll
        if file.find('PEIS')!=-1:
            cell_name = os.path.basename(file).split("_", 1)[0]
            break

    if eis == True:
        'Finding correct EIS files and formatting'
        dta_files = [file for file in os.listdir(folder_loc) if file.endswith('.DTA')] #Makes a list of all .DTA files in the folder loc
        bias_deg_peis = [] #initializing bias files list
        
        if a10 == False: #Plot the EIS for the first 10 hours in 1 hr increments (the increment it is taken in)
            for file in dta_files: #Finding all fuel cell bias EIS files
                if (file.find('PEIS')!=-1) and (file.find('_Deg')!=-1) and (file.find('Deg10')==-1) and (file.find('n3Bias')!=-1):
                    bias_deg_peis.append(os.path.join(folder_loc,file))

            f10_bias_deg_peis = sorted(bias_deg_peis, key=lambda x: int((x[x.find('__#')+len('__#'):x.rfind('.DTA')]))) #Sorts numerically by time
            
            'Plotting EIS'
            for peis in f10_bias_deg_peis:
                # --- Finding and formatting
                loc = os.path.join(folder_loc, peis) # creates the full path to the file
                time = peis[peis.find('__#')+len('__#'):peis.rfind('.DTA')] #gets the bias from the file name
                
                nyquist_name =  str(time) + ' Hours'

                # --- Plotting
                plot_peiss(area,nyquist_name,loc,legend_loc=legend_loc)
            
            plt.show()

        if a10 == True: # Plot the EIS after 10 hours in 10 hour increments (the increment it is taken in)
            
            'Finding and formatting'
            for file in dta_files: #Finding all fuel cell bias EIS files
                if (file.find('PEIS_n3Bias')!=-1) and (file.find('_Deg10')!=-1):# and (file.find('n3Bias')!=-1):
                    loc = os.path.join(folder_loc,file)
                    bias_deg_peis.append((loc,int(fl.get_timestamp(loc).strftime("%s"))))

            'Sorting values by time'
            bias_deg_peis.sort(key=lambda x:x[1])
            a10_bias_deg_peis = bias_deg_peis

            # --- Getting the ent time
            last = a10_bias_deg_peis[-1][1]
            end_time = int((last-t0)/3600) #hrs

            'Setting up an array for the color map'
            cmap = plt.cm.get_cmap('plasma',end_time)

            'Plotting EIS'
            fig,ax = plt.subplots()
            for peis in a10_bias_deg_peis:
                loc = os.path.join(folder_loc, peis[0]) # creates the full path to the file

                # --- Finding time of the EIS from the start of the degradation test
                test_time = peis[1]
                time = round((test_time-t0)/3600) #hrs
                nyquist_name =  str(round(time)) + ' Hours'

                # --- Plotting
                df_useful = peis_data(area,loc)
                if cbar == True:
                    ax.plot(df_useful['ohm'],df_useful['ohm.1'],'o',markersize=9,label = nyquist_name,color = cmap(time)) #plots data
                if cbar == False:
                    plot_peiss(area,nyquist_name,loc,ncol=ncol,legend_loc=legend_loc,color = cmap(time))

                # --- Plot Formatting                  
                ax.set_xlabel('Zreal (\u03A9 cm$^2$)',fontsize='xx-large')
                ax.set_ylabel('$\u2212$Zimag (\u03A9 cm$^2$)',fontsize='xx-large')
                ax.tick_params(axis='both', which='major', labelsize='x-large')
                ax.axhline(y=0,color='k', linestyle='-.') # plots line at 0 #D2492A is the color of Mines orange
                ax.axis('scaled') #Keeps X and Y axis scaled 1 to 1
                ax.spines['top'].set_visible(False)
                ax.spines['right'].set_visible(False)

            if cbar == True:
                sm = plt.cm.ScalarMappable(cmap=cmap)
                sm.set_array([0,end_time])
                divider = make_axes_locatable(ax)
                cax = divider.append_axes("right", size="5%", pad=0.05)
                cb = plt.colorbar(sm,ticks = [0,end_time],cax=cax)
                cb.set_label(label='Time (hrs)',fontsize = 'xx-large',labelpad = -10)
                cb.set_ticks(ticks = [0,end_time],fontsize='xx-large')
                cb.ax.tick_params(labelsize='x-large')

            plt.tight_layout()
            plt.show()

    if drt == True:
        'Finding and Formatting'
        fig, ax = plt.subplots() #initializing plots for DRT
        
        if a10 == False: #Plot the DRT for the first 10 hours in 1 hr increments (the increment it is taken in)
            f10_bias_deg_fits = [file for file in os.listdir(jar_loc) if (file.find('n3_Bias')!=-1 and file.find('n3_Bias10')==-1)] #Makes a list of all Bias files 
            f10_bias_deg_fits = sorted(f10_bias_deg_fits, key=lambda x: int((x[x.find('__#')+len('__#'):x.rfind('.pkl')]))) #Sorts numerically by time
            
            for fit in f10_bias_deg_fits: # Calling fits and plotting
                # --- Finding and formatting
                loc = os.path.join(folder_loc, fit) # creates the full path to the file
                time = fit[fit.find('__#')+len('__#'):fit.rfind('.pkl')] #gets the bias from the file name
                map_fit_name = cell_name+'_n3_Bias__#' + time + '.pkl'
                label = str(time) + ' Hours'

                # --- Plotting
                inv = Inverter()
                inv.load_fit_data(os.path.join(jar_loc,map_fit_name))
                bp.plot_distribution(None,inv,ax,unit_scale='',label = label)
            
            ax.legend()
            plt.show()
        
        if a10 == True:  
            'Sorting out relevant DRT files'         
            a10_bias_fits = [file for file in os.listdir(jar_loc) if file.find('n3_Bias10')!=-1] #Makes a list of all Bias files 
            a10_bias_fits = natsort.humansorted(a10_bias_fits)

            "Finding the time from the correct EIS file"
            dta_files = [file for file in os.listdir(folder_loc) if file.endswith('.DTA')] #Makes a list of all .DTA files in the folder loc
            bias_deg_peis_time = [] #initializing bias files list
            
            for file in dta_files: #Finding all fuel cell bias EIS files
                if (file.find('PEIS')!=-1) and (file.find('_Deg10')!=-1) and (file.find('n3Bias')!=-1):
                    loc = os.path.join(folder_loc,file)
                    bias_deg_peis_time.append(int(fl.get_timestamp(loc).strftime("%s")))
            
            'Sorting values by time'
            a10_bias_deg_peis_time = natsort.humansorted(bias_deg_peis_time)

            'Setting up an array for the color map'
            cmap = plt.cm.get_cmap('plasma')
            color_space = np.linspace(0,1,len(a10_bias_fits)) # array from 0-1 for the colormap for plotting
            c = 0 # index of the color array

            i = 0 # Setting the initial index
            for fit in a10_bias_fits: #For loop to plot the DRT of all PO2 files (probably a more elegant way, but this works)
                # --- Finding the right mapfit file and matching it to the right time
                number = fit[fit.find('_Bias10')+len('_Bias10'):fit.rfind('.pkl')] #gets the bias from the file name
                map_fit_name = cell_name + '_n3_Bias10' + number + '.pkl'
                time = round((a10_bias_deg_peis_time[i]-t0)/3600) # to convert to hours and round to the nearest hour from the test start
                label = str(time) + ' Hours'

                i = i+1

                # --- Plotting
                inv = Inverter()
                inv.load_fit_data(os.path.join(jar_loc,map_fit_name))
                color = cmap(color_space[c])
                bp.plot_distribution(None,inv,ax,unit_scale='',label = label,color=color)
                c = c + 1

            ax.legend()
            plt.show()

def ECstb_ocv_eis_plots(folder_loc:str, jar_loc:str, area:float, first_file:str = 'default',
                        a10:bool = True, eis:bool = True, drt:bool = True, ncol:int = 1,
                         legend_loc:str = 'outside',cbar = True):
    '''
    Plots the EIS and the fit DRT of a cell taken at OCV during a long term stability test. This function
    complements a Gamry sequence that I use to test the cell stability over time in electrolysis cell (EC) mode
    DRT must already be fit before using this function.

    OCV eis spectra are taken one an hour for the first 10 hours then 1 every 10 hours for the rest of the test

    Parameters
    ----------
    folder_loc, str: (path to a directory)
        The folder location of the EIS files 
    jar_loc, str: (path to a directory)
        The folder location of the pickle jar where the DRT fits are 
    area, float: 
        The active cell area in cm^2
    start_file, str: (default = 'default')
        The name of the first file taken in the stability sequence
    a10, bool: (After 10 hours) (default = True)
        Whether or not to plot the first 10 hours of the test.
        if True, then only hours 10 onward are plotted
    eis, bool: (default = True)
        Whether or not to plot the EIS spectra
    drt, bool: (default = True)
        Whether or not to plot the DRT fits
    ncol, int: (default = 1)
        The number of columns to use in the legend of the plot
    legend_loc, str:  (default = 'outside')
        The location of the legend (default = 'best'). The other option is to put 
        the legend outside the figure and this is done by setting legend_loc = 'outside'
    cbar, Bool: (default = True)
        If this is set to true then the legend will just be a colorbar
        If false, then the normal legend will exist

    Return --> None but one or more plots are created and shown
    '''

    # cell_name = os.path.basename(folder_loc).split("_", 3)[2]

    if first_file == 'default': #if another file is specified as the first file, this file will be used to find T0
        for file in os.listdir(folder_loc): #Finding the first file
            if file.find('ECstability__#1.DTA')!=-1 and file.find('OCV')!=-1:
                start_file = os.path.join(folder_loc,file)
        t0 = int(fl.get_timestamp(start_file).strftime("%s")) # Getting time stamp for first file in s from epoch, and convert to int
    else:
        t0 = int(fl.get_timestamp(start_file).strftime("%s")) # Getting time stamp for first file in s from epoch, and convert to int

    for file in os.listdir(folder_loc): # Finding the name of the celll
        if file.find('PEIS')!=-1:
            cell_name = os.path.basename(file).split("_", 1)[0]
            break

    if eis == True:
        'Finding correct EIS files and formatting'
        dta_files = [file for file in os.listdir(folder_loc) if file.endswith('.DTA')] #Makes a list of all .DTA files in the folder loc
        ocv_deg_peis = [] #initializing bias files list
        
        if a10 == False: #Plot the EIS for the first 10 hours in 1 hr increments (the increment it is taken in)
            for file in dta_files: #Finding all fuel cell bias EIS files
                if (file.find('PEIS')!=-1) and (file.find('_ECstability')!=-1) and (file.find('ECstability10')==-1) and (file.find('TNV')==-1):
                    ocv_deg_peis.append(os.path.join(folder_loc,file))

            f10_ocv_deg_peis = sorted(ocv_deg_peis, key=lambda x: int((x[x.find('__#')+len('__#'):x.rfind('.DTA')]))) #Sorts numerically by time
            
            'Plotting EIS'
            for peis in f10_ocv_deg_peis:
                # --- Finding and formatting
                loc = os.path.join(folder_loc, peis) # creates the full path to the file
                time = peis[peis.find('__#')+len('__#'):peis.rfind('.DTA')] #gets the bias from the file name
                
                nyquist_name =  str(time) + ' Hours'

                # --- Plotting
                plot_peiss(area,nyquist_name,loc,ncol=ncol,legend_loc=legend_loc)
            plt.show()

        if a10 == True: # Plot the DRT after 10 hours in 10 hour increments (the increment it is taken in)
            'Finding and formatting'
            for file in dta_files: #Finding all fuel cell bias EIS files
                if (file.find('PEIS')!=-1) and (file.find('_ECstability10')!=-1) and (file.find('TNV')==-1):
                    loc = os.path.join(folder_loc,file)
                    ocv_deg_peis.append((loc,int(fl.get_timestamp(loc).strftime("%s"))))
            
            'Sorting values by time'
            ocv_deg_peis.sort(key=lambda x:x[1])
            a10_ocv_deg_peis = ocv_deg_peis

            # --- Getting the ent time
            last = a10_ocv_deg_peis[-1][1]
            end_time = int((last-t0)/3600) #hrs

            'Setting up an array for the color map'
            color = np.linspace(0,1,len(a10_ocv_deg_peis)) # array from 0-1 for the colormap for plotting
            cmap = plt.cm.get_cmap('cmr.neon', end_time)

            'Plotting EIS'
            fig,ax = plt.subplots()
            for peis in a10_ocv_deg_peis:
                loc = os.path.join(folder_loc, peis[0]) # creates the full path to the file

                # --- Finding time of the EIS from the start of the degradation test
                test_time = peis[1]
                time = int((test_time-t0)/3600) #hrs
                nyquist_name =  str(time) + ' Hours'

                # --- Plotting
                # plot_peiss(area,nyquist_name,loc,ncol=ncol,legend_loc=legend_loc,color = cmap(color[c]))
                df_useful = peis_data(area,loc)
                if cbar == True:
                    ax.plot(df_useful['ohm'],df_useful['ohm.1'],'o',markersize=9,label = nyquist_name,color = cmap(time)) #plots data
                if cbar == False:
                    plot_peiss(area,nyquist_name,loc,ncol=ncol,legend_loc=legend_loc,color = cmap(time))

                # --- Plot Formatting                  
                ax.set_xlabel('Zreal (\u03A9 cm$^2$)',fontsize='xx-large')
                ax.set_ylabel('$\u2212$Zimag (\u03A9 cm$^2$)',fontsize='xx-large')
                ax.tick_params(axis='both', which='major', labelsize='x-large')
                ax.axhline(y=0,color='k', linestyle='-.') # plots line at 0 #D2492A is the color of Mines orange
                ax.axis('scaled') #Keeps X and Y axis scaled 1 to 1
                ax.spines['top'].set_visible(False)
                ax.spines['right'].set_visible(False)

            if cbar == True:
                sm = plt.cm.ScalarMappable(cmap=cmap)
                sm.set_array([0,end_time])
                divider = make_axes_locatable(ax)
                cax = divider.append_axes("right", size="5%", pad=0.05)
                cb = plt.colorbar(sm,ticks = [0,end_time],cax=cax)
                cb.set_label(label='Time (hrs)',fontsize = 'xx-large',labelpad = -16)
                cb.set_ticks(ticks = [0,end_time],fontsize='xx-large')
                cb.ax.tick_params(labelsize='x-large')

            plt.tight_layout()
            plt.show()

    if drt == True:
        'Finding and Formatting'
        fig, ax = plt.subplots() #initializing plots for DRT

        if a10 == False: #Plot the DRT for the first 10 hours in 1 hr increments (the increment it is taken in)
            f10_ocv_deg_fits = [file for file in os.listdir(jar_loc) if (file.find('OCV')!=-1 and file.find('OCV10')==-1)] #Makes a list of all OCV files 
            f10_ocv_deg_fits = sorted(f10_ocv_deg_fits, key=lambda x: int((x[x.find('__#')+len('__#'):x.rfind('.pkl')]))) #Sorts numerically by time
            
            for fit in f10_ocv_deg_fits: # Calling fits and plotting
                # --- Finding and formatting
                loc = os.path.join(folder_loc, fit) # creates the full path to the file
                time = fit[fit.find('__#')+len('__#'):fit.rfind('.pkl')] #gets the bias from the file name
                map_fit_name = cell_name +'_ECstb_OCV__#'+time+'.pkl'
                label = str(time) + ' Hours'

                # --- Plotting
                inv = Inverter()
                inv.load_fit_data(os.path.join(jar_loc,map_fit_name))
                bp.plot_distribution(None,inv,ax,unit_scale='',label = label)

            ax.legend()
            plt.show()
        
        if a10 == True:  
            'Sorting out relevant DRT files'         
            a10_ocv_fits = [file for file in os.listdir(jar_loc) if file.find('OCV10')!=-1] #Makes a list of all OCV files 
            a10_ocv_fits = natsort.humansorted(a10_ocv_fits)

            "Finding the time from the correct EIS file"
            dta_files = [file for file in os.listdir(folder_loc) if file.endswith('.DTA')] #Makes a list of all .DTA files in the folder loc
            ocv_deg_peis_time = [] #initializing bias files list
            for file in dta_files: #Finding all fuel cell ocv EIS files
                if (file.find('PEIS')!=-1) and (file.find('_ECstability10')!=-1) and (file.find('TNV')==-1):
                    loc = os.path.join(folder_loc,file)
                    ocv_deg_peis_time.append(fl.get_timestamp(loc).strftime("%s"))

            'Sorting values by time'
            # The order of the list is reversed to ensure that if there is degradation, the DRT plot can fit the larger spectra
            a10_ocv_deg_peis_time = natsort.humansorted(ocv_deg_peis_time)
            a10_ocv_deg_peis_time.sort(reverse=True)
            

            'Setting up an array for the color map'
            cmap = plt.cm.get_cmap('cmr.neon')
            color_space = np.linspace(0,1,len(a10_ocv_fits)) # array from 0-1 for the colormap for plotting
            c = 0 # index of the color array

            i = 0 # Setting the initial index
            for fit in reversed(a10_ocv_fits): #For loop to plot the DRT of all PO2 files (probably a more elegant way, but this works)
                # --- Finding the right mapfit file and matching it to the right time
                number = fit[fit.find('OCV10')+len('OCV10'):fit.rfind('.pkl')] #gets the OCV (which is actually a string)
                map_fit_name = cell_name +'_ECstb_OCV10'+number+'.pkl'
                time = round((int(a10_ocv_deg_peis_time[i])-t0)/3600) # to convert to hours and round to the nearest hour from the test start
                label = str(time) + ' Hours'
                i = i+1
                # print(number,', ',map_fit_name, ', ', time, ', ',label)

                # --- Plotting
                inv = Inverter()
                inv.load_fit_data(os.path.join(jar_loc,map_fit_name))
                color = cmap(color_space[len(color_space)-c-1])
                bp.plot_distribution(None,inv,ax,unit_scale='',label = label,color=color)
                c = c + 1

            ax.legend()
            plt.show()

def ECstb_bias_eis_plots(folder_loc:str, jar_loc:str, area:float, first_file:str = 'default',
                        a10:bool = True, eis:bool = True, drt:bool = True, cutoff_inductance = True,
                        ncol:int = 1, legend_loc:str = 'outside', cbar:bool = True, rev = True):
    '''
    Plots the EIS and the fit DRT of a cell taken at bias during a long term stability test. This function
    complements a Gamry sequence that I use to test the cell stability over time in electrolysis cell (EC) mode. 
    DRT must already be fit before using this function.

    Bias eis spectra are taken one an hour for the first 10 hours then 1 every 10 hours for the rest of the test.
    The tests are done at around the thermo-neutral voltage of the cell.
    
    Parameters
    ----------
    folder_loc, str: (path to a directory)
        The folder location of the EIS files
    jar_loc, str:  (path to a directory)
        The folder location of the pickle jar where the DRT fits are stored
    area, float: 
        The active cell area in cm^2
    start_file, str: (default = 'default')
        The name of the first file taken in the stability sequence. If 'default' is selected, then
        the function will find the first ocv file taken at the start of the stability sequence
    a10, bool: (After 10 hours) (default = True)
        Whether or not to plot the first 10 hours of the test.
        if True, then only hours 10 onward are plotted
    eis, bool: (default = True)
        Whether or not to plot the EIS spectra
    drt, bool: (default = True)
        Whether or not to plot the DRT fits
    cutoff_inductance, bool: (default = True)
        Graphically cuts off the inductance by setting the ylim to -0.05. If the cell is high performing
        the inductance takes up a huge area on the graph and the EIS spectra is hard to see.
    ncol, int: (default` = 1)
        The number of columns to use in the legend of the plot
    legend_loc, str: (default = 'outside')
        The location of the legend (default = 'best'). The other option is to put 
        the legend outside the figure and this is done by setting legend_loc = 'outside'
    cbar, Bool: (default = True)
        If this is set to true then the legend will just be a colorbar
        If false, then the normal legend will exist
    rev, Bol: (Default = True)
        Reverses the order that the DRT is shown in. This needs to be a variable because
        a10_ocv_fits and a10_ocv_deg_peis_time need to move in tandem or the legend will be off

    Return --> None but one or more plots are created and shown
    '''

    # cell_name = os.path.basename(folder_loc).split("_", 3)[2]

    if first_file == 'default': #if another file is specified as the first file, this file will be used to find T0
        for file in os.listdir(folder_loc): #Finding the first file
            if file.find('ECstability__#1.DTA')!=-1 and file.find('OCV')!=-1:
                start_file = os.path.join(folder_loc,file)
        t0 = int(fl.get_timestamp(start_file).strftime("%s")) # Getting time stamp for first file in s from epoch, and convert to int
    else:
        t0 = int(fl.get_timestamp(start_file).strftime("%s")) # Getting time stamp for first file in s from epoch, and convert to int

    for file in os.listdir(folder_loc): # Finding the name of the celll
        if file.find('PEIS')!=-1:
            cell_name = os.path.basename(file).split("_", 1)[0]
            break

    if eis == True:
        'Finding correct EIS files and formatting'
        dta_files = [file for file in os.listdir(folder_loc) if file.endswith('.DTA')] #Makes a list of all .DTA files in the folder loc
        bias_deg_peis = [] #initializing bias files list
        
        if a10 == False: #Plot the EIS for the first 10 hours in 1 hr increments (the increment it is taken in)
            for file in dta_files: #Finding all fuel cell bias EIS files
                if (file.find('PEIS_TNV')!=-1) and (file.find('_ECstability')!=-1) and (file.find('ECstability10')==-1):
                    bias_deg_peis.append(os.path.join(folder_loc,file))

            f10_bias_deg_peis = sorted(bias_deg_peis, key=lambda x: int((x[x.find('__#')+len('__#'):x.rfind('.DTA')]))) #Sorts numerically by time
            
            'Plotting EIS'
            for peis in f10_bias_deg_peis:
                # --- Finding and formatting
                loc = os.path.join(folder_loc, peis) # creates the full path to the file
                time = peis[peis.find('__#')+len('__#'):peis.rfind('.DTA')] #gets the bias from the file name
                
                nyquist_name =  str(time) + ' Hours'

                # --- Plotting
                plot_peiss(area,nyquist_name,loc,legend_loc=legend_loc)

                if cutoff_inductance == True:
                    plt.gca().set_ylim(bottom=-0.05)

            plt.show()

        if a10 == True: # Plot the EIS after 10 hours in 10 hour increments (the increment it is taken in)
            'Finding and formatting'
            for file in dta_files: #Finding all fuel cell bias EIS files
                if (file.find('PEIS_TNV')!=-1) and (file.find('_ECstability10')!=-1):
                    loc = os.path.join(folder_loc,file)
                    bias_deg_peis.append((loc,int(fl.get_timestamp(loc).strftime("%s"))))

            'Sorting values by time'
            bias_deg_peis.sort(key=lambda x:x[1])
            a10_bias_deg_peis = bias_deg_peis

            # --- Getting the ent time
            last = a10_bias_deg_peis[-1][1]
            end_time = int((last-t0)/3600) #hrs

            'Setting up an array for the color map'
            color = np.linspace(0,1,len(a10_bias_deg_peis)) # array from 0-1 for the colormap for plotting
            # c = 0 # index of the color array
            cmap = plt.cm.get_cmap('cividis',end_time)

            'Plotting EIS'
            fig,ax = plt.subplots()
            for peis in a10_bias_deg_peis:
                loc = os.path.join(folder_loc, peis[0]) # creates the full path to the file

                # --- Finding time of the EIS from the start of the degradation test
                test_time = peis[1]
                time = int((test_time-t0)/3600) #hrs
                nyquist_name =  str(round(time)) + ' Hours'

                # --- Plotting
                # plot_peiss(area,nyquist_name,loc,ncol=ncol,legend_loc=legend_loc,color = cmap(color[c]))
                df_useful = peis_data(area,loc)
                if cbar == True:
                    ax.plot(df_useful['ohm'],df_useful['ohm.1'],'o',markersize=9,label = nyquist_name,color = cmap(time)) #plots data
                if cbar == False:
                    plot_peiss(area,nyquist_name,loc,ncol=ncol,legend_loc=legend_loc,color = cmap(time))
                
                if cutoff_inductance == True:
                    plt.gca().set_ylim(bottom=-0.05)

                # --- Plot Formatting                  
                ax.set_xlabel('Zreal (\u03A9 cm$^2$)',fontsize='xx-large')
                ax.set_ylabel('$\u2212$Zimag (\u03A9 cm$^2$)',fontsize='xx-large')
                ax.tick_params(axis='both', which='major', labelsize='x-large')
                ax.axhline(y=0,color='k', linestyle='-.') # plots line at 0 #D2492A is the color of Mines orange
                ax.axis('scaled') #Keeps X and Y axis scaled 1 to 1
                ax.spines['top'].set_visible(False)
                ax.spines['right'].set_visible(False)
            
                # c = c + 1
            
            if cbar == True:
                sm = plt.cm.ScalarMappable(cmap=cmap)
                sm.set_array([0,end_time])
                divider = make_axes_locatable(ax)
                cax = divider.append_axes("right", size="5%", pad=0.05)
                cb = plt.colorbar(sm,ticks = [0,end_time],cax=cax)
                cb.set_label(label='Time (hrs)',fontsize = 'xx-large',labelpad = -16)
                cb.set_ticks(ticks = [0,end_time],fontsize='xx-large')
                cb.ax.tick_params(labelsize='x-large')

            plt.tight_layout()
            plt.show()

    if drt == True:
        'Finding and Formatting'
        fig, ax = plt.subplots() #initializing plots for DRT
        
        if a10 == False: #Plot the DRT for the first 10 hours in 1 hr increments (the increment it is taken in)
            f10_bias_deg_fits = [file for file in os.listdir(jar_loc) if (file.find('TNV')!=-1 and file.find('TNV10')==-1)] #Makes a list of all Bias files 
            f10_bias_deg_fits = sorted(f10_bias_deg_fits, key=lambda x: int((x[x.find('__#')+len('__#'):x.rfind('.pkl')]))) #Sorts numerically by time
            
            for fit in f10_bias_deg_fits: # Calling fits and plotting
                # --- Finding and formatting
                loc = os.path.join(folder_loc, fit) # creates the full path to the file
                time = fit[fit.find('__#')+len('__#'):fit.rfind('.pkl')] #gets the bias from the file name
                map_fit_name = cell_name+'_ECstb_TNV__#' + time + '.pkl'
                label = str(time) + ' Hours'

                # --- Plotting
                inv = Inverter()
                inv.load_fit_data(os.path.join(jar_loc,map_fit_name))
                bp.plot_distribution(None,inv,ax,unit_scale='',label = label)

            ax.legend()
            plt.show()
        
        if a10 == True:  
            'Sorting out relevant DRT files'         
            a10_bias_fits = [file for file in os.listdir(jar_loc) if file.find('TNV10')!=-1] #Makes a list of all OCV files 
            a10_bias_fits = natsort.humansorted(a10_bias_fits,reverse = rev)

            "Finding the time from the correct EIS file"
            dta_files = [file for file in os.listdir(folder_loc) if file.endswith('.DTA')] #Makes a list of all .DTA files in the folder loc
            bias_deg_peis_time = [] #initializing bias files list
            
            for file in dta_files: #Finding all fuel cell bias EIS files
                if (file.find('PEIS')!=-1) and (file.find('_ECstability10')!=-1) and (file.find('TNV')!=-1):
                    loc = os.path.join(folder_loc,file)
                    bias_deg_peis_time.append(int(fl.get_timestamp(loc).strftime("%s")))
            
            'Sorting values by time'
            a10_bias_deg_peis_time = natsort.humansorted(bias_deg_peis_time,reverse = rev)

            'Setting up an array for the color map'
            if rev == True:
                cmap = plt.cm.get_cmap('cividis_r')
            if rev == False:
                cmap = plt.cm.get_cmap('cividis')

            color_space = np.linspace(0,1,len(a10_bias_fits)) # array from 0-1 for the colormap for plotting
            c = 0 # index of the color array

            i = 0 # Setting the initial index
            for fit in a10_bias_fits: #For loop to plot the DRT of all PO2 files (probably a more elegant way, but this works)
                # --- Finding the right mapfit file and matching it to the right time
                number = fit[fit.find('_TNV10')+len('_TNV10'):fit.rfind('.pkl')] #gets the bias from the file name
                map_fit_name = cell_name + '_ECstb_TNV10' + number + '.pkl'
                time = round((a10_bias_deg_peis_time[i]-t0)/3600) # to convert to hours and round to the nearest hour from the test start
                label = str(time) + ' Hours'
                i = i+1

                # --- Plotting
                inv = Inverter()
                inv.load_fit_data(os.path.join(jar_loc,map_fit_name))
                color = cmap(color_space[c])
                bp.plot_distribution(None,inv,ax,unit_scale='',label = label,color=color)
                c = c + 1

            ax.legend()
            plt.show()

def excel_datasheet_exists(excel_file:str,sheet_name:str):
    '''
    Finds if the excel file exists and if the sheet name exists in the excel file
    If the excel file does not exist, it will return a writer that can create excel files
    If the excel file does exist, it will return a writer that can append to the excel file
    Will return whether or not the sheet_name exists in the excel file

    param excel_file,str: The excel file to look for
    param sheet_name,str: The name of the excel sheet to look for

    Return --> appropriate writer, whether or not the excel sheet exists
    '''
    
    exists = False
    if os.path.exists(excel_file)==True: # Looking for the data excel file in the button cell folder
        writer = pd.ExcelWriter(excel_file,engine='openpyxl',mode='a') #Creates a Pandas Excel writer using openpyxl as the engine in append mode
        wb = load_workbook(excel_file, read_only=True) # Looking for the bias Deg eis
        if sheet_name in wb.sheetnames:
            exists = True
    elif os.path.exists(excel_file)==False:
        writer = pd.ExcelWriter(excel_file,engine='xlsxwriter')
    
    return exists,writer

def o2_drt_peaks(folder_loc:str, tau_low:float, tau_high:float, concs:np.array = None,
                        rmv_concs_r:np.array = None, rmv_concs_l:np.array = None):
    '''
    This function is meant to linearly fit a single DRT peak across an oxygen concentration range
    This function can only be used after the po2_plots function
    This function plots ln(1/asr) vs. ln(O2 concentration (%))

    Parameters
    ----------
    folder_loc, str: (path to a directory)
        The location of the directory containing the EIS files.
    tau_low, float:
        The lower bound of the time constant range to fit.
    tau_high, float:
        The upper bound of the time constant range to fit.
    concs, np.array: (default = None)
        The concentrations that the DRT peaks are taken at. if this is none all temperature ranges in
        the cell data sheet will be fit.
    rmv_concs_r, np.array: (default = None)
        If two clusters overlap, specify the concentrations where there are overlap and this will remove
        the peaks with higher time constants (lower frequency, to the right) from the fit.
    rmv_concs_l, np.array: (default = None)
        If two clusters overlap, specify the concentrations where there are overlap and this will remove
        the peaks with lower time constants (higher frequency, to the left) from the fit.

    Returns --> The slope of ln(1/asr)/ln(O2%), and a plot of the DRT peak fit and the activation energy is calculated and printed on the plot
    '''

    # ----- Sorting out the points from a specific time constant
    for file in os.listdir(folder_loc):
        if file.endswith('Data.xlsx'):
            data_file = os.path.join(folder_loc,file)
            break
    
    data = pd.read_excel(data_file,'pO2 DRT peak fits')
    data = data[(data['Tau']>tau_low) & (data['Tau']<tau_high)]

    # ----- If desired, only plotting certain temperatures
    if concs is not None:
        data = data[data['O2 Concentration (%)'].isin(concs)]

    # ----- removing a duplicate if there is overlap between clusters to the right (remove points to the right)
    if rmv_concs_r is not None:
        for conc in rmv_concs_r:
            # - find the lowest tau for a given temperature
            conc_data = data[data['O2 Concentration (%)']==conc]
            conc_data = conc_data.sort_values(by='Tau')

            try:
                highest_tau = conc_data.iloc[-1]['Tau']

            except IndexError as error:
                traceback.print_exc()
                print('The removed concentration must be in concs array (aka concentrations plotted)')
                print('Check the concs array and the rmv_concs array')
                print('If that does not work, make sure that there are data points in the specified tau range (it is easy to confuse the log scale)')
                sys.exit(1)

            # - remove all higher tau values
            data = data[data['Tau']!=highest_tau]

    # ----- removing a duplicate if there is overlap between clusters to the left (remove points to the left)
    if rmv_concs_l is not None:
        for conc in rmv_concs_l:
            # - find the lowest tau for a given temperature
            conc_data = data[data['O2 Concentration (%)']==conc]
            conc_data = conc_data.sort_values(by='Tau')

            try:
                lowest_tau = conc_data.iloc[0]['Tau']

            except IndexError as error:
                traceback.print_exc()
                print('The removed concentration must be in concs array (aka concentrations plotted)')
                print('Check the concs array and the rmv_concs array')
                print('If that does not work, make sure that there are data points in the specified tau range (it is easy to confuse the log scale)')
                sys.exit(1)

            # - remove all higher tau values
            data = data[data['Tau']!=lowest_tau]
    
    # ----- Finding low and high Tau values of the values plotted
    min_tau = data['Tau'].min()
    max_tau = data['Tau'].max()

    # ----- Formatting data for plotting
    rp_asr = data['Resistance'].values
    concs = data['O2 Concentration (%)'].values
    ln_rp_asr = np.log(1/rp_asr)
    
    o2_concs = np.array(concs)
    ln_o2 = np.log(o2_concs/100)

    # ----- Plotting
    x = ln_o2
    y = ln_rp_asr

    fig, ax1 = plt.subplots()
    ax2 = ax1.twiny()
    ax1.plot(x,y,'ko')

    # - Setting font sizes
    label_fontsize = 'x-large'
    tick_fontsize = 'large'
    text_fontsize = 'x-large'

    # - Setting ticks
    ax1.set_xticks(x, np.round(x,2), fontsize=tick_fontsize)
    ax1.set_yticks(y,labels= np.round(y,2), fontsize=tick_fontsize)
    ax2.set_xticks(x, labels = concs, fontsize=tick_fontsize)
    ax1.set_xlim(-1.7,0.1)
    ax2.set_xlim(-1.7,0.1)

    # - linear Fit:
    mr,br = np.polyfit(ln_o2,ln_rp_asr,1)
    mr,br, r, p_value, std_err = scipy.stats.linregress(x, y)
    fit_r = mr*ln_o2 + br
    ax1.plot(x, fit_r,'g')

    # - Axis labels:
    ax1.set_xlabel('ln(O$_2$) (%)',fontsize=label_fontsize)
    ax2.set_xlabel('Oxygen Concentration (%)',fontsize=label_fontsize)
    ax1.set_ylabel('ln(1/ASR) (S/cm$^2$)',fontsize=label_fontsize)

    # - Creating table:
    row_labels = ['r squared']
    slopes = f'{round(mr,2)}'
    table_values = [[round(r**2,3)]] # $\\bf{round(mr,2)}$
    color = 'palegreen'

    if mr >= 0:
        table = plt.table(cellText=table_values,colWidths = [.2]*3,rowLabels=row_labels,loc = 'lower right',
            rowColours= [color,color])
    else:
        table = plt.table(cellText=table_values,colWidths = [.2]*3,rowLabels=row_labels,loc = 'lower center',
            rowColours= [color,color])

    table.scale(1,1.6)
    
    # - Printing figure text
    mr_str = f'{round(mr,2)}'
    tau_lows = f'{min_tau:.2e}'
    tau_highs = f'{max_tau:.2e}'
    if mr >=0:
        fig.text(0.68,0.22,r'ASR$_\mathrm{P}$ Slope = '+mr_str,weight='bold',fontsize = tick_fontsize)
        fig.text(0.17,0.81,'DRT peak between '+tau_lows+'(\u03C4/s) and '+tau_highs+'(\u03C4/s)',fontsize = tick_fontsize)
    else:
        fig.text(0.37,0.22,r'ASR$_\mathrm{P}$ Slope = '+mr_str,weight='bold',fontsize = tick_fontsize)
        fig.text(0.38,0.81,'DRT peak between '+tau_lows+'(\u03C4/s) and '+tau_highs+'(\u03C4/s)',fontsize = tick_fontsize)
    plt.tight_layout()

    plt.show()

    return(mr)







